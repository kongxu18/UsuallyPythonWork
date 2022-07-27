import json
import sys
import io
import openpyxl
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import json
import sys
import os
from openpyxl.styles import Border,Side
from openpyxl.styles import PatternFill
import 绘制索分段数据
from openpyxl.worksheet.pagebreak import Break

def GenBorder(left_style, right_style, top_style, bottom_style):
    rltBorder = Border()
    if left_style != None:
        rltBorder.left = Side(border_style=left_style,color='000000')
    if right_style != None:
        rltBorder.right = Side(border_style=right_style,color='000000')
    if top_style != None:
        rltBorder.top = Side(border_style=top_style,color='000000')
    if bottom_style != None:
        rltBorder.bottom = Side(border_style=bottom_style,color='000000')
    return rltBorder

thinSide = Side(border_style='thin',color='000000')
mediumSide = Side(border_style='medium',color='000000')
dashedSide = Side(border_style='dashed',color='000000')
doubleSide = Side(border_style='double',color='000000')
mediumDashedSide = Side(border_style='mediumDashed',color='000000')

titleBorder = GenBorder('thin','thin','medium','medium')
titleBorder_left = GenBorder('thin','dashed','medium','medium')
titleBorder_right = GenBorder('dashed','thin','medium','medium')
titleBorder_mid = GenBorder('dashed','dashed','medium','medium')

headerBorder = Border(left=Side(border_style='medium',color='000000'),
right=Side(border_style='medium',color='000000'),
top=Side(border_style='medium',color='000000'),
bottom=Side(border_style='medium',color='000000'))

mBorder_LeftRight = Border(left=Side(border_style='medium',color='000000'),
right=Side(border_style='medium',color='000000'),)

mtBorder_LeftRight = Border(left=Side(border_style='medium',color='000000'),
right=Side(border_style='thin',color='000000'),)

thBorder_LeftRight = Border(left=Side(border_style='thin',color='000000'),
right=Side(border_style='thin',color='000000'),)

tmBorder_LeftRight = Border(left=Side(border_style='thin',color='000000'),
right=Side(border_style='medium',color='000000'),)


rowFill = PatternFill(fill_type = 'solid',start_color='EEEEEE',end_color='000000')

层号 = 2

downDoubleBorder = Border(bottom=Side(border_style='double',color='000000'))

excelPath = "D:\\rhinoWork\\F%d索网数据明细.xlsx"%(层号)
jsonPath = "F%d索网数据.json"%(层号)

with open(jsonPath, 'r', encoding='utf-8') as f:
    json_data=json.loads(f.read())
wb = Workbook()
wb.remove(wb.active)
colWith = 18
titleFontStyle = Font(size = "20")

def createSheet(索网名称):
    sheet = wb.create_sheet(索网名称)
    sheet.merge_cells('A1:G1')
    sheet['A1'] = "%s轴索长明细"%(索网名称)
    sheet['A1'].alignment = Alignment(vertical="center")
    sheet['A1'].font = titleFontStyle
    for c in range(0,7):
        sheet[1][c].border = titleBorder
    sheet.column_dimensions["A"].width = colWith
    sheet.column_dimensions["B"].width = colWith
    sheet.column_dimensions["C"].width = colWith
    sheet.column_dimensions["D"].width = colWith
    sheet.column_dimensions["E"].width = colWith
    sheet.column_dimensions["F"].width = colWith
    sheet.column_dimensions["G"].width = colWith
    return sheet

headerUseFont = Font(name='宋体', size=20)
contentUseFont = Font(name='宋体', size=14)
titleUseFont = Font(name='宋体', size=12, bold=True)

def initHeaderCell(cell):
    cell.border = GenBorder('thin','thin', 'thin', 'thin')
    cell.font = headerUseFont
    cell.alignment = Alignment(vertical="center",horizontal="center")

def initContentCell(cell):
    cell.border = GenBorder('thin','thin', 'thin', 'thin')
    cell.font = contentUseFont
    cell.alignment = Alignment(vertical="center",horizontal="center")

def initTitleCell(cell):
    cell.border = GenBorder('thin','thin', 'thin', 'thin')
    cell.fill = titleFill
    cell.font = titleUseFont
    cell.alignment = Alignment(vertical="center",horizontal="center")

def 写入索段数据(sheet,索段数据,baseRow):
    #数据写入
    rowIndex = baseRow
    initContentCell(sheet.cell(row=rowIndex, column=1, value='序号'))
    initContentCell(sheet.cell(row=rowIndex, column=2, value='索号'))
    initContentCell(sheet.cell(row=rowIndex, column=3, value='索径'))
    initContentCell(sheet.cell(row=rowIndex, column=4, value='轴线号'))
    initContentCell(sheet.cell(row=rowIndex, column=5, value='无应力索长mm'))
    initContentCell(sheet.cell(row=rowIndex, column=6, value='预张拉索长mm'))
    initContentCell(sheet.cell(row=rowIndex, column=7, value='标记张力kn'))
    sheet.row_dimensions[rowIndex].height = 20
    rowIndex += 1
    initContentCell(sheet.cell(row=rowIndex, column=1, value=索段数据['序号']))
    initContentCell(sheet.cell(row=rowIndex, column=2, value=索段数据['编号']))
    initContentCell(sheet.cell(row=rowIndex, column=3, value='φ%d'%(索段数据['索头规格'])))
    initContentCell(sheet.cell(row=rowIndex, column=4, value=索段数据['轴线号']))
    initContentCell(sheet.cell(row=rowIndex, column=5, value='%.2f'%(索段数据['无应力长度'])))
    initContentCell(sheet.cell(row=rowIndex, column=6, value='%.2f'%(索段数据['预张拉长度'])))
    initContentCell(sheet.cell(row=rowIndex, column=7, value=索段数据['Fi']))
    sheet.row_dimensions[rowIndex].height = 20
    rowIndex += 1
    imgPath = 绘制索分段数据.绘制索段(索段数据)
    img = openpyxl.drawing.image.Image(imgPath)
    img.anchor = 'A%d'%(rowIndex)
    sheet.add_image(img)
    imgColCount = 8
    for row in range(rowIndex,rowIndex + imgColCount):
        sheet.row_dimensions[row].height = 25
    return rowIndex + imgColCount


index = 0
count索网data = len(json_data['data'])
for 索网dataI,索网data in enumerate(json_data['data']):
    index += 1
    if index == 2:
        break
    useSheet = createSheet(索网data['编号'])
    row_index = 2
    print ('%d/%d'%(索网dataI,count索网data))
    索段_arr = 索网data['索段_arr']
    for 索段data in 索段_arr:
        row_index = 写入索段数据(useSheet,索段data,row_index)
    # next_page_horizon, next_page_vertical = sheet.page_breaks
    # if 索网dataI >= 0:
    #     break

# wb.save(excelPath)
