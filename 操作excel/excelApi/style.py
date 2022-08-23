from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors

border_style = Border(left=Side(border_style='thin', color='000000'),
                      right=Side(border_style='thin', color='000000'),
                      top=Side(border_style='thin', color='000000'),
                      bottom=Side(border_style='thin', color='000000'))

border_style2 = Border(left=Side(border_style='thin', color='000000'),
                       right=Side(border_style='thin', color='000000'),
                       top=Side(border_style='thin', color='000000'),
                       bottom=Side(border_style='thin', color='000000'),
                       diagonal=Side(border_style='thin', color='000000'),
                       diagonalUp=False, diagonalDown=True)

alignment_center = Alignment(horizontal='center', vertical='center', wrapText=True)
alignment_left = Alignment(horizontal='left', vertical='center', wrapText=True)
alignment_justify = Alignment(horizontal='justify', vertical='justify', wrap_text=True)
fill_red = PatternFill(fill_type="solid", fgColor="EE292B")
fill_green = PatternFill(fill_type="solid", fgColor="7FCA40")
fill_yellow = PatternFill(fill_type="solid", fgColor='EAC121')
fill_gray = PatternFill(fill_type="solid", fgColor='A6A6A6')
fill_blue = PatternFill(fill_type="solid", fgColor='8DB4E2')
fill_220 = PatternFill(fill_type='solid',fgColor='FFC0CB')
fill_珊瑚红 = PatternFill(fill_type='solid',fgColor='FF7F00')
fill_石板蓝 = PatternFill(fill_type='solid',fgColor='007FFF')
fill_长石色 = PatternFill(fill_type='solid',fgColor='D19275')
fill_新棕褐色 = PatternFill(fill_type='solid',fgColor='EBC791')
fill_石板蓝 = PatternFill(fill_type='solid',fgColor='ADEAEA')
fill_亮天蓝色 = PatternFill(fill_type='solid',fgColor='38B0DE')

fill_浅蓝 = PatternFill(fill_type='solid',fgColor='B8CCE4')
fill_浅粉 = PatternFill(fill_type='solid',fgColor='F2DCDB')

fill_棕黄色 = PatternFill(fill_type='solid',fgColor='F79646')



font12 = Font(name='宋体', bold=False, size=12)
font14 = Font(name='宋体', bold=False, size=14)
font16 = Font(name='宋体', bold=False, size=16)
font8 = Font(name='宋体', bold=False, size=8)
font10 = Font(name='宋体', bold=False, size=10)


font12b = Font(name='宋体', bold=True, size=12)
font14b = Font(name='宋体', bold=True, size=14)
font16b = Font(name='宋体', bold=True, size=16)
font8b = Font(name='宋体', bold=True, size=8)


