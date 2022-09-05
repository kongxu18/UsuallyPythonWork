import pandas as pd

from data import SqlData
from style import *
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import typing


class Title:
    def __init__(self, s_row=1, s_col=1):
        self.s_row = s_row
        self.s_col = s_col
        self.__content = None
        self.sheet = None  # type:Workbook()
        self.width = None

    @property
    def content(self):
        return self.__content

    @content.setter
    def content(self, val):
        if isinstance(val, str):
            self.__content = val
        else:
            raise ValueError('title must be str type')

    @staticmethod
    def set_style(cell):
        cell.font = font16b
        cell.border = border_style
        # cell.quotePrefix = False
        cell.alignment = alignment_center

    def merge(self):
        self.sheet.merge_cells(start_row=self.s_row, start_column=self.s_col,
                               end_row=self.s_row, end_column=self.s_col + self.width - 1)

    def add(self):
        if self.sheet is not None:
            if self.content:
                cell = self.sheet.cell(row=self.s_row, column=self.s_col, value=self.content)
                self.set_style(cell)
                self.merge()
            else:
                raise ValueError('标题不能为空')
        return self.width


class Header(object):
    def __init__(self, s_row=2, s_col=1):
        self.s_row = s_row
        self.s_col = s_col
        self.__content = None
        self.sheet = None  # type:Workbook()

    @property
    def content(self):
        return self.__content

    @content.setter
    def content(self, val):
        if isinstance(val, list):
            self.__content = val
        else:
            raise ValueError('header must be list type')

    @staticmethod
    def set_style(cell):
        cell.font = font12b
        cell.border = border_style
        # cell.quotePrefix = False
        cell.alignment = alignment_center

    def add(self):
        if self.content is not None:
            for i, val in enumerate(self.content):
                cell = self.sheet.cell(row=self.s_row, column=self.s_col + i, value=val)
                self.set_style(cell)


class Body(object):
    def __init__(self, s_row=3, s_col=1):
        self.s_row = s_row
        self.s_col = s_col
        self.__dataframe = None
        self.sheet = None  # type:Workbook()
        self.__model = None

    @property
    def dataframe(self):
        return self.__dataframe

    @dataframe.setter
    def dataframe(self, df):
        if isinstance(df, pd.DataFrame):
            self.__dataframe = df
        else:
            raise ValueError('body must be dataframe type')

    @staticmethod
    def set_style(cell, left=False, right=False, fill=False):
        cell.font = font12
        cell.border = border_style
        if left:
            cell.border = border_style_left_dashed
        if right:
            cell.border = border_style_right_dashed
        if fill:
            cell.fill = fill_浅蓝
        # cell.quotePrefix = False
        cell.alignment = alignment_center

    def add(self):
        if self.dataframe is not None:
            height, width = self.dataframe.shape

            for i, row in self.dataframe.iterrows():
                for w in range(width - 1):
                    val = row[w]

                    cell = self.sheet.cell(row=self.s_row + i, column=self.s_col + w)

                    if val:
                        cell.value = val

                    # 特殊设置虚线
                    if row['isfill'] == 1:
                        self.set_style(cell, fill=True)
                    elif w == 6 or w == 8:
                        self.set_style(cell, right=True)
                    elif w == 7 or w == 9:
                        self.set_style(cell, left=True)
                    else:
                        self.set_style(cell)
            return height


class Excel:
    member = ['title', 'header', 'body']

    def __init__(self, builder):
        self.path = builder.path
        self.header = builder.header
        self.wb = builder.wb
        self.sheet = builder.sheet  # type:Workbook().active
        self.height = getattr(builder, 'height')
        self.width = builder.width

    def set_row_height(self):
        for row in range(1, self.height + 1):
            if row == 1:
                height = 40
            elif row == 2:
                height = 25
            else:
                height = 17
            self.sheet.row_dimensions[row].height = height

    @staticmethod
    def isChinese(word):
        count = 0
        if not word:
            return 0
        word = str(word)
        try:
            for ch in word:
                if '\u4e00' <= ch <= '\u9fff':
                    count += 2 * 1.5
                else:
                    count += 1 * 1.5
        except Exception as err:
            print(word, '--ischinese')
        return count

    def set_row_width(self):
        header_content = self.header.content
        widths = [10, 8, 8, 20, 15, 15, 15, 15, 15, 15]
        for i, name in enumerate(header_content):
            letter = get_column_letter(i + 1)
            self.sheet.column_dimensions[letter].width = widths[i]

    def freeze_panes(self):
        self.sheet.freeze_panes = 'A3'

    def create(self):
        self.freeze_panes()
        self.set_row_height()
        self.set_row_width()
        self.wb.save(self.path)


class ExcelBuilder(object):
    def __init__(self, path):
        self.path = path
        self.wb = self.wb_create()
        self._sheet = None
        self.header = None
        self.width = 0
        self.height = 0
        self.excel = None

    @staticmethod
    def wb_create():
        return Workbook()

    @property
    def sheet(self):
        if self._sheet is None:
            self._sheet = self.wb.active
        return self._sheet

    @sheet.setter
    def sheet(self, name):
        self._sheet = self.wb.create_sheet(name)

    def add_title(self, arg):
        self.height += 1
        self.setter(self.sheet, arg)

    def add_header(self, arg):
        self.height += 1
        self.header = arg
        self.setter(self.sheet, arg)

    def add_body(self, arg):
        rows = self.setter(self.sheet, arg)
        self.height += rows

    def build(self):
        self.excel = Excel(self)
        return self.excel.create()

    @staticmethod
    def setter(sheet, arg):
        setattr(arg, 'sheet', sheet)
        if hasattr(arg, 'add'):
            res = arg.add()
            return res


class ModelOne(object):
    def __init__(self, df):
        self.df = df
        self.LH = None
        self.LM1 = None
        self.LM2 = None
        self.LM3 = None
        self.LM4 = None
        self.LM5 = None
        self.res = pd.DataFrame(columns=['单体名称', '列号', '行号', '构件全称', '铝条编号',
                                         '下斜长度', '左角度', '右角度', '左切口长度', '右切口长度', 'isfill', 'row'])

    def group_name(self):
        return self.df.groupby('构件全称')

    @staticmethod
    def front_part(df):
        单体名称 = df['全局名称'].iloc[0]
        列号 = df['列号'].iloc[0]
        行号 = df['行号'].iloc[0]
        return 单体名称, 列号, 行号

    @property
    def dataframe(self):

        for name, df in self.group_name():
            front_part = self.front_part(df)
            back_part = self.deal_needed(df)
            self.set_all(df, back_part, name, front_part)

        res = self.res.sort_values(by=['单体名称', '列号', '行号', '构件全称', 'row']).reset_index(drop=True)
        return res.iloc[:, :11]

    @staticmethod
    def deal_needed(df):
        d_li = list(df[df['参数描述'] == 'LM零件总长'].sort_values(by=['参数序号']).loc[:, '参数值'])
        b_li = list(df[df['参数描述'] == 'LM端口切口长'].sort_values(by=['参数序号']).loc[:, '参数值'])
        p_li = list(df[df['参数描述'] == 'LM端口角度'].sort_values(by=['参数序号']).loc[:, '参数值'])
        return d_li, b_li, p_li

    def set_all(self, df, back_part, name, front_part):
        单体名称, 列号, 行号 = front_part
        D, B, P = back_part
        if len(D) != 6:
            raise ValueError('%s：D参数 个数不为6' % name)
        if len(B) != 8 or len(P) != 8:
            raise ValueError('%s:B,P 参数个数不为8' % name)

        df = df[df['参数名称'] == 'L1']
        if len(df) != 1:
            raise ValueError('L1 个数不为1')
        val = df['参数值'].iloc[0]

        table = [
            [单体名称, 列号, 行号, name, 'LH', val, None, None, None, None, 1, 1],
            [单体名称, 列号, 行号, name, 'LM1', D[0], P[0], None, B[0], None, 0, 2],
            [单体名称, 列号, 行号, name, 'LM2', D[1], None, P[1], None, B[1], 0, 3],
            [单体名称, 列号, 行号, name, 'LM3', D[2], P[2], P[3], B[2], B[3], 0, 4],
            [单体名称, 列号, 行号, name, 'LM4', D[3], P[4], None, B[4], None, 0, 5],
            [单体名称, 列号, 行号, name, 'LM5', D[4], None, P[5], None, B[5], 0, 6],
            [单体名称, 列号, 行号, name, 'LM6', D[5], P[6], P[7], B[6], B[7], 0, 7]
        ]
        df_table = pd.DataFrame(data=table, columns=self.res.columns)
        self.res = pd.concat([self.res, df_table], axis=0)


if __name__ == '__main__':
    sql = 'select * from FT254E铝框架零件参数(329) order by 列号'

    sql_res = SqlData(sql=sql).data
    # print(d)
    model1 = ModelOne(sql_res)
    #
    bodyData = model1.dataframe
    body = Body()
    body.dataframe = bodyData

    title = Title()
    title.content = '使用的铝条型号请参考零件图图纸'
    title.width = 10

    header = Header()
    header.content = ['单体名称', '列号', '行号', '构件全称', '铝条编号', '下料长度', '左角度', '右角度', '左切口长度', '右切口长度']

    excel = ExcelBuilder('test.xlsx')

    excel.add_title(title)
    excel.add_header(header)
    excel.add_body(body)

    excel.build()
