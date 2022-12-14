import json
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
import string

alignment_center = Alignment(horizontal='center', vertical='center')


class JsonData:
    def __init__(self, path: str, encode='utf8'):
        self.path = path
        self.encoding = encode

    @property
    def data(self) -> dict:
        with open(self.path, 'r', encoding=self.encoding) as f:
            res_data = json.loads(f.read()).get('json')
        return res_data


class BaseData(object):
    keyword = ''
    start_row = ''
    letter = string.ascii_uppercase

    def __init__(self, json_data: dict):
        self.json_data = json_data
        self.header_area = ''

    @property
    def data(self):
        return self.json_data.get(self.keyword)

    @property
    def length(self):
        return len(self.data)

    @staticmethod
    def get_letter_coord(col, row):
        col -= 1
        row -= 1
        row = str(row + 1)
        if col < 26:
            index = col + ord('A')
            return chr(index), row
        else:
            col_1 = (col // 26) - 1
            col_2 = (col % 26)
            return BaseData.letter[col_1] + BaseData.letter[col_2], row


class Header(BaseData):
    def __init__(self, json_data: dict):
        super().__init__(json_data)


class Rows(BaseData):
    def __init__(self, json_data: dict):
        super().__init__(json_data)


class Title(BaseData):
    def __init__(self, json_data: dict):
        super().__init__(json_data)


class HeaderSettings(BaseData):
    def __init__(self, json_data: dict):
        super().__init__(json_data)


class Picture():
    def __init__(self, w, h, path):
        self.h = w
        self.w = h
        self.path = path
        self.col = None
        self.row = None


class DataFactory(object):
    font = None
    type = ''
    start_row = ''
    keyword = ''

    def __init__(self, DataClass, *args, **kwargs):
        self.DataClass = DataClass
        self.args = args
        self.kwargs = kwargs
        self.pic_dict = None

    def create(self):
        dataIns = self.DataClass(*self.args, **self.kwargs)
        dataIns.keyword = self.keyword
        if self.keyword == 'headerSetting':
            return dataIns
        dataIns.font = self.font
        dataIns.start_row = self.start_row

        return dataIns


class HeaderFactory(DataFactory):

    def __init__(self, DataClass, *args, **kwargs):
        super().__init__(DataClass, *args, **kwargs)
        self.font = Font(name='??????', bold=True)
        self.start_row = 2
        self.keyword = 'headers'


class TitleFactory(DataFactory):
    def __init__(self, DataClass, *args, **kwargs):
        super().__init__(DataClass, *args, **kwargs)
        self.font = Font(name='??????', bold=True)
        self.start_row = 1
        self.keyword = 'title'


class RowsFactory(DataFactory):
    def __init__(self, DataClass, *args, **kwargs):
        super().__init__(DataClass, *args, **kwargs)
        self.font = Font(name='??????', bold=True)
        self.start_row = 3
        self.keyword = 'rows'


class HeaderSettingFactory(DataFactory):
    def __init__(self, DataClass, *args, **kwargs):
        super().__init__(DataClass, *args, **kwargs)
        self.keyword = 'headerSetting'


class PictureFactory(DataFactory):
    def __init__(self, DataClass, *args, **kwargs):
        super().__init__(DataClass, *args, **kwargs)
        self.keyword = 'picture'
        self.font = None
        self.type = 'picture'

    def picture(self, pic):
        ...


class Excel:
    member = ['title', 'header', 'rows', 'picture']
    autoIndex = None
    addQuet = 0

    def __init__(self, builder):
        self.numb_hs = {}
        self.pict_hs = {}
        self.text_hs = {}
        self.path = builder.path
        self.title = builder.title
        self.header = builder.header
        self.rows = builder.rows
        self.wb = builder.wb
        self.sheet = builder.sheet
        self.headerSettings = None

    def deal_headerSetting(self):

        header_list: list = self.header.data
        if not self.headerSettings:
            return None
        for setting_dict in self.headerSettings:
            name = setting_dict.get('name')
            index = header_list.index(name)
            header_type = setting_dict.get('type')
            if header_type == '??????':
                self.text_hs.update({index: setting_dict})
            elif header_type == '??????':
                self.pict_hs.update({index: setting_dict})
            elif header_type == '??????':
                self.numb_hs.update({index: setting_dict})

    @property
    def pict_setting_dict(self) -> dict:
        return self.pict_hs

    @property
    def text_setting_dict(self) -> dict:
        return self.text_hs

    @property
    def numb_setting_dict(self) -> dict:
        return self.numb_hs

    @staticmethod
    def get_letter_coord(col, row):
        col -= 1
        row -= 1
        row = str(row + 1)
        if col < 26:
            index = col + ord('A')
            return chr(index), row
        else:
            col_1 = (col // 26) - 1
            col_2 = (col % 26)
            return BaseData.letter[col_1] + BaseData.letter[col_2], row

    def set_col_width(self):
        start_col = 2 if self.autoIndex else 1

        for key, val in self.text_setting_dict.items():
            col_letter, __ = self.get_letter_coord(start_col + int(key), 1)
            self.sheet.column_dimensions[col_letter].width = int(val.get('width'))

    def create(self):
        self.set_col_width()
        self.wb.save(self.path)


class ExcelBuilder(object):

    def __init__(self, title, header, rows, path='test.xlsx'):
        self.title = title
        self.header = header
        self.rows = rows
        self.path = path
        self.wb = Workbook()
        self.sheet = self.wb.active
        self.excel = Excel(self)
        self.autoIndex = self.excel.autoIndex
        self.addQuet = self.excel.addQuet

    def set_header_settings(self, headerSettings):
        # ???????????????
        if not isinstance(headerSettings, list):
            headerSettings = headerSettings.data
        self.excel.headerSettings = headerSettings
        self.excel.deal_headerSetting()

    def add_title(self):
        length = self.header.length
        if self.autoIndex:
            length += 1
        start_row = self.title.start_row
        start_col = 1
        cell = self.sheet.cell(row=start_row, column=start_col, value=self.title.data)
        # ???????????????
        self.sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=length)
        cell.font = self.header.font
        cell.quotePrefix = 1 if self.addQuet else 0
        cell.alignment = alignment_center

    def add_header(self):
        start_row = self.header.start_row
        start_col = 2 if self.autoIndex else 1
        for i, val in enumerate(self.header.data):
            cell = self.sheet.cell(row=start_row, column=start_col + i, value=val)
            cell.font = self.header.font
            cell.quotePrefix = 1 if self.addQuet else 0
            cell.alignment = alignment_center

    def add_rows(self):
        # ??????????????????
        pict_setting_dict = self.excel.pict_setting_dict

        start_row = self.rows.start_row
        start_col = 2 if self.autoIndex else 1

        for row, obj in enumerate(self.rows.data):
            self.rows_height_list = []
            self.rows_exist_pic = None
            for col in range(self.header.length):
                header_name = self.header.data[col]
                cell_val = obj.get(header_name)
                # ??????
                serial_number = self.sheet.cell(row=start_row + row, column=1, value=row + 1)

                # ?????????????????????????????????
                if pict_setting_dict.get(col) and cell_val:
                    self.add_picture(self.rows, row + start_row, col + start_col, cell_val, pict_setting_dict.get(col))
                    self.rows_exist_pic = True
                    cell_val = None

                cell = self.sheet.cell(row=start_row + row, column=start_col + col, value=cell_val)
                # ??????
                cell.alignment = alignment_center
                serial_number.alignment = alignment_center
            if self.rows_exist_pic:
                self.sheet.row_dimensions[row + start_row].height = max(self.rows_height_list)

    def add_picture(self, rows_obj, row, col, img_path, setting_dict: dict):
        picWidth = float(setting_dict.get('picWidth'))
        picHeight = float(setting_dict.get('picHeight'))

        if getattr(rows_obj, 'get_letter_coord', None):
            handle = getattr(rows_obj, 'get_letter_coord', None)
            colLetter, rowIndex = handle(col, row)

            img = Image(img_path)

            img.width = picWidth
            img.height = picHeight

            cell_width, cell_height = self.covert(picWidth, picHeight)
            self.rows_height_list.append(cell_height)
            self.sheet.column_dimensions[colLetter].width = cell_width
            self.sheet.add_image(img, colLetter + rowIndex)

    def covert(self, img_width: float, img_height: float):
        """
        100???????????? = 2.62cm
        1cm = 28.3465??? = 4.374 ?????????

        :return: ???????????????????????????????????? ???
        """

        w_cm = (2.62 / 100) * img_width

        w_strL = w_cm * 6
        # ???
        h_cm = (2.62 / 100) * img_height

        h_libra = h_cm * 28.3465 + 10
        return w_strL, h_libra

    def build(self):
        return self.excel


if __name__ == '__main__':
    data = JsonData('test1.json').data

    title = TitleFactory(Title, data).create()
    header = HeaderFactory(Header, data).create()

    rows = RowsFactory(Rows, data).create()

    header_setting = HeaderSettingFactory(HeaderSettings, data).create()

    excel_builder = ExcelBuilder(title, header, rows, 'test.xlsx')
    # # ??????settings
    excel_builder.set_header_settings(header_setting)
    #
    ExcelBuilder.autoIndex = 1
    excel_builder.add_title()
    excel_builder.add_header()
    excel_builder.add_rows()
    excel = excel_builder.build()
    excel.create()
