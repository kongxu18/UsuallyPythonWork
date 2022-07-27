import json
from openpyxl.styles import Font, Border, Side
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
import string
import io
import sys
import os.path
import typing

border_style = Border(left=Side(border_style='thin', color='000000'),
                      right=Side(border_style='thin', color='000000'),
                      top=Side(border_style='thin', color='000000'),
                      bottom=Side(border_style='thin', color='000000'))

alignment_center = Alignment(horizontal='center', vertical='center',wrapText=True)


class JsonData:
    def __init__(self, path: str, encode='utf8'):
        self.path = path
        self.encoding = encode

    @property
    def data(self) -> dict:
        with open(self.path, 'r', encoding=self.encoding) as f:
            res_data = json.loads(f.read())
        return res_data


class BaseData(object):
    keyword = ''
    start_row = ''
    letter = string.ascii_uppercase

    def __init__(self, json_data: dict):
        self.json_data = json_data
        self.header_area = ''

    @property
    def data(self):
        return self.json_data.get(self.keyword)

    @property
    def length(self):
        return len(self.data)

    @staticmethod
    def get_letter_coord(col, row):
        col -= 1
        row -= 1
        row = str(row + 1)
        if col < 26:
            index = col + ord('A')
            return chr(index), row
        else:
            col_1 = (col // 26) - 1
            col_2 = (col % 26)
            return BaseData.letter[col_1] + BaseData.letter[col_2], row


class Header(BaseData):
    def __init__(self, json_data: dict):
        super().__init__(json_data)


class Rows(BaseData):
    def __init__(self, json_data: dict):
        super().__init__(json_data)


class Title(BaseData):
    def __init__(self, json_data: dict):
        super().__init__(json_data)


class HeaderSettings(BaseData):
    def __init__(self, json_data: dict):
        super().__init__(json_data)


class Picture():
    def __init__(self, w, h, path):
        self.h = w
        self.w = h
        self.path = path
        self.col = None
        self.row = None


class DataFactory(object):
    font = None
    type = ''
    start_row = ''
    keyword = ''
    border = ''

    def __init__(self, DataClass, *args, **kwargs):
        self.DataClass = DataClass
        self.args = args
        self.kwargs = kwargs
        self.pic_dict = None

    def create(self):
        dataIns = self.DataClass(*self.args, **self.kwargs)
        dataIns.keyword = self.keyword
        if self.keyword == 'headerSetting':
            return dataIns
        dataIns.font = self.font
        dataIns.start_row = self.start_row
        dataIns.border = self.border

        return dataIns


class HeaderFactory(DataFactory):

    def __init__(self, DataClass, *args, **kwargs):
        super().__init__(DataClass, *args, **kwargs)
        self.font = Font(name='黑体', bold=True)
        self.border = border_style
        self.start_row = 2
        self.keyword = 'headers'


class TitleFactory(DataFactory):
    def __init__(self, DataClass, *args, **kwargs):
        super().__init__(DataClass, *args, **kwargs)
        self.border = border_style
        self.font = Font(name='黑体', bold=True, size=20)
        self.start_row = 1
        self.keyword = 'title'


class RowsFactory(DataFactory):
    def __init__(self, DataClass, *args, **kwargs):
        super().__init__(DataClass, *args, **kwargs)
        self.border = border_style
        self.font = Font(name='黑体', bold=True)
        self.start_row = 3
        self.keyword = 'rows'


class HeaderSettingFactory(DataFactory):
    def __init__(self, DataClass, *args, **kwargs):
        super().__init__(DataClass, *args, **kwargs)
        self.keyword = 'headerSetting'


class PictureFactory(DataFactory):
    def __init__(self, DataClass, *args, **kwargs):
        super().__init__(DataClass, *args, **kwargs)
        self.keyword = 'picture'
        self.font = None
        self.type = 'picture'

    def picture(self, pic):
        ...


class Excel:
    member = ['title', 'header', 'rows', 'picture']
    autoIndex = None
    addQuet = 0

    def __init__(self, builder):
        self.numb_hs = {}
        self.pict_hs = {}
        self.text_hs = {}
        self.path = builder.path
        self.title = builder.title
        self.header = builder.header
        self.rows = builder.rows
        self.wb = builder.wb
        self.sheet = builder.sheet  # type:Workbook().active
        self.headerSettings = None

    def deal_headerSetting(self):
        header_list = self.header.data
        if not self.headerSettings:
            return None
        for setting_dict in self.headerSettings:
            name = setting_dict.get('name')
            index = header_list.index(name)
            header_type = setting_dict.get('type')
            if header_type == '文本':
                self.text_hs.update({index: setting_dict})
            elif header_type == '图片':
                self.pict_hs.update({index: setting_dict})
            elif header_type == '数字':
                self.numb_hs.update({index: setting_dict})

    @property
    def pict_setting_dict(self) -> dict:
        return self.pict_hs

    @property
    def text_setting_dict(self) -> dict:
        return self.text_hs

    @property
    def numb_setting_dict(self) -> dict:
        return self.numb_hs

    @staticmethod
    def get_letter_coord(col, row):
        col -= 1
        row -= 1
        row = str(row + 1)
        if col < 26:
            index = col + ord('A')
            return chr(index), row
        else:
            col_1 = (col // 26) - 1
            col_2 = (col % 26)
            return BaseData.letter[col_1] + BaseData.letter[col_2], row

    def set_col_width(self):
        start_col = 2 if self.autoIndex else 1

        for key, val in self.text_setting_dict.items():
            col_letter, __ = self.get_letter_coord(start_col + int(key), 1)
            self.sheet.column_dimensions[col_letter].width = int(val.get('width'))

    def set_print_area(self):
        max_column, max_row = self.get_letter_coord(self.sheet.max_column, self.sheet.max_row)

        self.sheet.print_area = "A1:%s%d" % (max_column, int(max_row))

    def set_oddHeader(self):
        title = self.title.data
        self.sheet.oddHeader.center.text = "%s" % (title)
        self.sheet.oddHeader.center.size = 16

    def set_oddFooter(self):
        self.sheet.oddFooter.center.text = '第&[Page]页 共&N页'

    def set_title_column_name(self):
        # 每一页打印标题
        col, __ = self.get_letter_coord(self.sheet.max_column, 2)

        self.sheet.print_title_cols = 'A:'+str(col)
        self.sheet.print_title_rows = '1:2'

    def create(self):
        self.set_col_width()
        self.set_print_area()
        # self.set_oddHeader()
        self.set_oddFooter()
        self.set_title_column_name()
        self.wb.save(self.path)


class ExcelBuilder(object):

    def __init__(self, title, header, rows, path='test.xlsx'):
        self.title = title
        self.header = header
        self.rows = rows
        self.path = path
        self.wb = Workbook()
        self.sheet = self.wb.active
        self.excel = Excel(self)
        self.autoIndex = self.excel.autoIndex
        self.addQuet = self.excel.addQuet

    def set_header_settings(self, headerSettings):
        # 传入的参数
        if not isinstance(headerSettings, list):
            headerSettings = headerSettings.data
        self.excel.headerSettings = headerSettings
        self.excel.deal_headerSetting()

    def add_title(self):
        length = self.header.length
        if self.autoIndex:
            length += 1
        start_row = self.title.start_row
        start_col = 1
        cell = self.sheet.cell(row=start_row, column=start_col, value=self.title.data)
        for i in range(length):
            cell_ = self.sheet.cell(row=start_row, column=i + start_col)
            cell_.border = self.title.border
        # 合并单元格
        self.sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=length)
        cell.font = self.title.font
        cell.border = self.title.border
        cell.quotePrefix = 1 if self.addQuet else 0
        cell.alignment = alignment_center
        self.sheet.row_dimensions[start_row].height = 25

    def add_header(self):
        start_row = self.header.start_row
        start_col = 1
        header_data = self.header.data
        if self.autoIndex:
            header_data = ['序号'] + self.header.data
        for i, val in enumerate(header_data):
            cell = self.sheet.cell(row=start_row, column=start_col + i, value=val)
            cell.font = self.header.font
            cell.border = self.header.border
            cell.quotePrefix = 1 if self.addQuet else 0
            cell.alignment = alignment_center
        self.sheet.row_dimensions[start_row].height = 22

    def add_rows(self):
        # 图片设置列表
        pict_setting_dict = self.excel.pict_setting_dict

        start_row = self.rows.start_row
        start_col = 2 if self.autoIndex else 1

        for row, obj in enumerate(self.rows.data):
            self.rows_height_list = []
            self.rows_exist_pic = None
            for col in range(self.header.length):
                header_name = self.header.data[col]
                cell_val = obj.get(header_name)

                # 判断这列是图片，并有值
                if pict_setting_dict.get(col) and cell_val:
                    cell_val = self.add_picture(self.rows, row + start_row, col + start_col, cell_val,
                                                pict_setting_dict.get(col))
                    self.rows_exist_pic = True

                cell = self.sheet.cell(row=start_row + row, column=start_col + col, value=cell_val)
                # 居中
                cell.alignment = alignment_center
                cell.border = self.rows.border

                # 序号
                if self.autoIndex:
                    serial_number = self.sheet.cell(row=start_row + row, column=1, value=row + 1)
                    serial_number.alignment = alignment_center
                    serial_number.border = self.rows.border

            if self.rows_exist_pic:
                self.sheet.row_dimensions[row + start_row].height = max(self.rows_height_list)
            else:
                self.sheet.row_dimensions[row + start_row].height = 18

    def add_picture(self, rows_obj, row, col, img_path, setting_dict: dict):
        picWidth = float(setting_dict.get('picWidth'))
        picHeight = float(setting_dict.get('picHeight'))

        if getattr(rows_obj, 'get_letter_coord', None):
            handle = getattr(rows_obj, 'get_letter_coord', None)
            colLetter, rowIndex = handle(col, row)
            img = None
            img_path = 'D:/website/TridentSystem/public' + img_path
            img_path = img_path.replace('/', '\\')
            if os.path.exists(img_path):
                img = Image(img_path)
                img.width = picWidth
                img.height = picHeight

            cell_width, cell_height = self.covert(picWidth, picHeight)
            self.rows_height_list.append(cell_height)
            self.sheet.column_dimensions[colLetter].width = cell_width

            if img:
                self.sheet.add_image(img, colLetter + rowIndex)
            else:
                return "图片未找到!"

    def covert(self, img_width: float, img_height: float):
        """
        100图片单位 = 2.62cm
        1cm = 28.3465镑 = 4.374 列单位

        :return: 单元格宽度字符长度，高度 镑
        """

        w_cm = (2.62 / 100) * img_width

        w_strL = w_cm * 4.8
        # 镑
        h_cm = (2.62 / 100) * img_height

        h_libra = h_cm * 28.3465 + 10
        return w_strL, h_libra

    def build(self):
        return self.excel


if __name__ == '__main__':
    # sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    # argv = sys.argv
    argv = ['', r'test.xlsx', r'ceshi.json', '1', '0']
    if len(argv) > 2:
        filePath = str(argv[1])
        jsonPath = str(argv[2])
        data = JsonData(jsonPath).data

        Excel.autoIndex = 0 if len(argv) <= 3 else (1 if argv[3] == '1' else 0)
        Excel.addQuet = 0 if len(argv) <= 4 else (1 if argv[4] == '1' else 0)
        title = TitleFactory(Title, data).create()

        header = HeaderFactory(Header, data).create()

        rows = RowsFactory(Rows, data).create()

        header_setting = HeaderSettingFactory(HeaderSettings, data).create()

        excel_builder = ExcelBuilder(title, header, rows, filePath)
        # 配置settings
        excel_builder.set_header_settings(header_setting)

        excel_builder.add_title()
        excel_builder.add_header()
        excel_builder.add_rows()
        excel = excel_builder.build()
        excel.create()
        print('OK', end='')
    else:
        print("No argv", end='')
