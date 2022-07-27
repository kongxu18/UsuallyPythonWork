import json
import sys
import io
import requests
from openpyxl import Workbook
from openpyxl.styles import Font
import json
import sys
import os

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
argv = sys.argv
#argv=['',r'd:\work\TridentSystem\filedata\excel\5fe9f712-301e-24dd-5233-4e0e74a138a7.xlsx',r'd:\work\TridentSystem\filedata\exceljson\5fe9f712-301e-24dd-5233-4e0e74a138a7.json','1','0']
if len(argv) > 2:
    filePath = str(argv[1])
    jsonPath = str(argv[2])
    autoIndex = False if len(argv) <= 3 else argv[3] == '1'
    addQuet = False if len(argv) <= 4 else argv[4] == '1'
    #print(addQuet)

    with open(jsonPath, 'r', encoding='utf-8') as f:
        retJson=json.loads(f.read())
    if not os.path.exists(os.path.dirname(filePath)):
        os.makedirs(os.path.dirname(filePath))
    wb = Workbook()
    sheet = wb.active
    headers = retJson['headers']
    rows = retJson['rows']
    titleFont = Font(name='黑体', bold=True)
    if(autoIndex):
        cell = sheet.cell(row=1, column=1, value="序号")
        cell.font=titleFont
        cell.quotePrefix=1 if addQuet else 0
    for index in range(0, len(headers)):
        #print(index, headers[index])
        cell = sheet.cell(row=1, column=index + (2 if autoIndex else 1), value=headers[index])
        cell.font=titleFont
        cell.quotePrefix=1 if addQuet else 0
    for rowIndex in range(0, len(rows)):
        rowjson = rows[rowIndex]
        if autoIndex:
            cell = sheet.cell(row=rowIndex + 2, column=1, value=rowIndex + 1)
            cell.quotePrefix=1 if addQuet else 0
        for colIndex in range(0, len(headers)):
            cellValue = rowjson[headers[colIndex]] if headers[colIndex] in rowjson.keys() else ''
            if cellValue is None:
                cellValue = ''
            cell = sheet.cell(row=rowIndex + 2, column=colIndex + (2 if autoIndex else 1), value=cellValue)
            cell.quotePrefix=1 if addQuet else 0
    wb.save(filePath)
    print('OK',end='')
else:
    print("No argv",end='')
