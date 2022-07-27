import json
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
import string

alignment_center = Alignment(horizontal='center', vertical='center')


class Excel:
    area = {}
    member = ['title', 'header', 'rows', 'picture']

    def __init__(self, builder, path):
        self.title = builder.title
        self.header = builder.header
        self.rows = builder.rows
        self.path = path
        self.wb = builder.wb

    def create(self):
        self.wb.save(self.path)


class ExcelBuilder():
    autoIndex = None
    addQuet = 0

    def __init__(self, title, header, rows):
        self.title = title
        self.header = header
        self.rows = rows
        self.wb = Workbook()
        self.sheet = self.wb.active

    def add_title(self):
        length = self.header.length
        if self.autoIndex:
            length += 1
        start_row = self.title.start_row
        start_col = 1
        cell = self.sheet.cell(row=start_row, column=start_col, value=self.title.data)
        # 合并单元格
        self.sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=length)
        cell.font = self.header.font
        cell.quotePrefix = 1 if self.addQuet else 0
        cell.alignment = alignment_center

    def add_header(self):

        start_row = self.header.start_row
        start_col = 2 if self.autoIndex else 1
        for i, val in enumerate(self.header.data):
            cell = self.sheet.cell(row=start_row, column=start_col + i, value=val)
            cell.font = self.header.font
            cell.quotePrefix = 1 if self.addQuet else 0
            cell.alignment = alignment_center

    def add_rows(self):

        start_row = self.rows.start_row
        start_col = 2 if self.autoIndex else 1
        pic_dict: dict = self.rows.pic_dict

        for row, obj in enumerate(self.rows.data):
            self.rows_height_list = []
            self.rows_exist_pic = None
            for col in range(self.header.length):
                header_name = self.header.data[col]
                cell_val = obj.get(header_name)
                # 判断这列是图片，并有值
                if pic_dict.get(col) and cell_val:
                    self.add_picture(self.rows, row + start_row, col + start_col, cell_val)
                    self.rows_exist_pic = True
                serial_number = self.sheet.cell(row=start_row + row, column=1, value=row + 1)
                cell = self.sheet.cell(row=start_row + row, column=start_col + col, value=cell_val)
                # 居中
                cell.alignment = alignment_center
                serial_number.alignment = alignment_center
            if self.rows_exist_pic:
                self.sheet.row_dimensions[row+start_row].height = max(self.rows_height_list)

    def add_picture(self, rows_obj, row, col, img_path):
        if getattr(rows_obj, 'get_letter_coord', None):
            handle = getattr(rows_obj, 'get_letter_coord', None)
            colLetter, rowIndex = handle(col, row)

            img = Image(img_path)

            img.width = 300
            img.height = 200
            cell_width, cell_height = self.covert(300, 200)
            self.rows_height_list.append(cell_height)
            self.sheet.column_dimensions[colLetter].width = cell_width
            self.sheet.add_image(img, colLetter + rowIndex)

    def covert(self, img_width, img_height):
        """
        100图片单位 = 2.62cm
        1cm = 28.3465镑 = 4.374 列单位

        :return: 单元格宽度字符长度，高度 镑
        """
        w_cm = 2.62 / 100 * img_width
        print(w_cm)
        w_strL = w_cm * 6
        # 镑
        h_cm = 2.62 / 100 * img_height
        print(h_cm)
        h_libra = h_cm * 28.3465 + 10
        return w_strL, h_libra

    def build(self, path):
        return Excel(self, path)


class JsonData:
    def __init__(self, path: str, encode='utf8'):
        self.path = path
        self.encoding = encode

    @property
    def data(self) -> dict:
        with open(self.path, 'r', encoding=self.encoding) as f:
            res_data = json.loads(f.read())
        return res_data


class BaseData(object):
    keyword = ''
    start_row = ''
    letter = string.ascii_uppercase

    def __init__(self, json_data: dict):
        self.json_data = json_data
        self.header_area = ''

    @property
    def data(self):
        return self.json_data.get(self.keyword)

    @property
    def area(self):
        return self.header_area

    @property
    def length(self):
        return len(self.data)

    @staticmethod
    def get_letter_coord(col, row):
        col -= 1
        row -= 1
        row = str(row + 1)
        if col < 26:
            index = col + ord('A')
            return chr(index), row
        else:
            col_1 = (col // 26) - 1
            col_2 = (col % 26)
            return BaseData.letter[col_1] + BaseData.letter[col_2], row


class Header(BaseData):
    pic_dict = {}

    def __init__(self, json_data: dict):
        super().__init__(json_data)

    @property
    def data(self):
        data_list = self.json_data.get(self.keyword)
        for i, item in enumerate(data_list):
            if isinstance(item, dict):
                data_list[i] = item.get('title')
                Header.pic_dict[i] = (int(item.get('w')), int(item.get('h')))
        return data_list


class Rows(BaseData):
    def __init__(self, json_data: dict):
        super().__init__(json_data)
        self.pic_dict = Header.pic_dict
        self.pic_list = []


class Title(BaseData):
    def __init__(self, json_data: dict):
        super().__init__(json_data)


class Picture():
    def __init__(self, w, h, path):
        self.h = w
        self.w = h
        self.path = path
        self.col = None
        self.row = None


class DataFactory(object):
    font = None
    type = ''
    start_row = ''
    keyword = ''

    def __init__(self, DataClass, *args, **kwargs):
        self.DataClass = DataClass
        self.args = args
        self.kwargs = kwargs
        self.pic_dict = None

    def create(self):
        dataIns = self.DataClass(*self.args, **self.kwargs)
        dataIns.font = self.font
        dataIns.keyword = self.keyword
        dataIns.start_row = self.start_row

        return dataIns


class HeaderFactory(DataFactory):

    def __init__(self, DataClass, *args, **kwargs):
        super().__init__(DataClass, *args, **kwargs)
        self.font = Font(name='黑体', bold=True)
        self.start_row = 2
        self.keyword = 'headers'


class TitleFactory(DataFactory):
    def __init__(self, DataClass, *args, **kwargs):
        super().__init__(DataClass, *args, **kwargs)
        self.font = Font(name='黑体', bold=True)
        self.start_row = 1
        self.keyword = 'title'


class RowsFactory(DataFactory):
    def __init__(self, DataClass, *args, **kwargs):
        super().__init__(DataClass, *args, **kwargs)
        self.font = Font(name='黑体', bold=True)
        self.start_row = 3
        self.keyword = 'rows'


class PictureFactory(DataFactory):
    def __init__(self, DataClass, *args, **kwargs):
        super().__init__(DataClass, *args, **kwargs)
        self.font = None

    def picture(self, pic):
        ...


if __name__ == '__main__':
    data = JsonData('t.json').data
    title = TitleFactory(Title, data).create()
    header = HeaderFactory(Header, data).create()
    rows = RowsFactory(Rows, data).create()

    excel_builder = ExcelBuilder(title, header, rows)
    ExcelBuilder.autoIndex = 1
    excel_builder.add_title()
    excel_builder.add_header()
    excel_builder.add_rows()
    excel = excel_builder.build('a.xlsx')
    excel.create()
