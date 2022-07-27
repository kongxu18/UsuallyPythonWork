import openpyxl

path = ['F1索网数据明细.xlsx', 'F2索网数据明细.xlsx', 'F3索网数据明细.xlsx',
        '变更索网数据明细20211008.xlsx', '变更索网数据明细20211015.xlsx', '变更索网数据明细20211018.xlsx']


class Excel(object):
    def __init__(self, path, name='_'):
        self.name = name
        self.path = path
        self.workbook = openpyxl.load_workbook(path)
        self.data = {'数据': []}

    def sheets(self):

        for sheet in self.workbook.worksheets:
            sheetName = sheet.title

            self.deal_cell(sheet, sheetName)

    def deal_cell(self, sheet, sheetName):
        """
        遍历
        :param sheet:
        :param data:
        :return:
        """

        sheet_data = {}
        _max_row = sheet.max_row
        _max_col = sheet.max_column
        for row in range(1, _max_row + 1):
            for col in range(1, _max_col + 1):
                cell = sheet.cell(row=row, column=col).value
                if cell:
                    if cell == '序号':
                        sheet_data['序号'] = sheet.cell(row=row + 1, column=col).value
                    elif cell == '索号':
                        sheet_data['索号'] = sheet.cell(row=row + 1, column=col).value
                    elif cell == '索径':
                        val: str = sheet.cell(row=row + 1, column=col).value
                        val = val.replace('φ', '')
                        sheet_data['索径'] = val
                    elif cell == '预张拉索长mm':
                        sheet_data['预张拉索长mm'] = sheet.cell(row=row + 1, column=col).value
                    elif cell == '累计长度':

                        arr = []
                        column = col
                        while True:
                            column += 1
                            length_cell = sheet.cell(row=row, column=column).value

                            length_cell = str(length_cell)
                            if length_cell == 'None' or length_cell is None:
                                break

                            arr.append(length_cell)

                        sheet_data['累计长度'] = arr
                        self.data['数据'].append(sheet_data)
                        sheet_data = {}

    def save(self):
        import json
        json_attr = json.dumps(self.data, ensure_ascii=False)
        with open(self.path[:4] + self.name + '.json', 'w', encoding='utf8') as f:
            f.write(json_attr)


if __name__ == '__main__':
    excel = Excel(path[0], '')
    excel.sheets()
    excel.save()

    excel2 = Excel(path[1], '')
    excel2.sheets()
    excel2.save()

    excel3 = Excel(path[2], '')
    excel3.sheets()
    excel3.save()

    excel4 = Excel(path[3], '1')
    excel4.sheets()
    excel4.save()

    excel5 = Excel(path[4], '2')
    excel5.sheets()
    excel5.save()

    excel6 = Excel(path[5], '3')
    excel6.sheets()
    excel6.save()
