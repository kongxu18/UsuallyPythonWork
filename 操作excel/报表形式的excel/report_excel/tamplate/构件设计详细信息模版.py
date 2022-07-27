import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.drawing.image import Image
import copy
import datetime

data = datetime.datetime.now()

border_style_1 = Border(bottom=Side(border_style='thick', color='000000'))
border_style_2 = Border(bottom=Side(border_style='thin', color='000000'))
font_1 = Font(name='黑体', bold=False, size=16)
alignment_center = Alignment(horizontal='center', vertical='center', wrapText=True)
wb = openpyxl.Workbook()

ws = wb.active
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.print_options.horizontalCentered = True
ws.page_margins.left = 0
ws.page_margins.right = 0

ws.print_area = 'A1:F10'
# ws.print_options.verticalCentered = True

ws.oddHeader.center.text = '上海海勃膜结构股份有限公司'
ws.oddHeader.center.font = '黑体,Bold'
ws.oddHeader.center.size = 24

ws.oddFooter.center.text = '第&[Page]页 共&N页'
ws.oddFooter.left.text = '打印人:'
ws.oddFooter.left.size = 10
ws.oddFooter.right.text = '打印时间:%s' % str(data)[:19]
ws.oddFooter.right.size = 10

cell_1 = ws.cell(row=1, column=1)
cell_1.value = '构件设计详细信息'
ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=6)
cell_1.alignment = alignment_center
cell_1.font = Font('黑体', size=16, bold=False)

cell_2 = ws.cell(row=2, column=1)
cell_2.value = '构件名称:DDJ(底座-东西轴-边跨段收边边角铁)'
ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=5)
cell_2.alignment = Alignment(horizontal='left', vertical='center')
cell_2.font = Font('黑体', size=14, bold=True)

cell_3 = ws.cell(row=2, column=6)
cell_3.value = '项目单体:G(体育馆)'
ws.merge_cells(start_row=2, end_row=2, start_column=6, end_column=6)
cell_3.alignment = Alignment(horizontal='left', vertical='center')
cell_3.font = Font('黑体', size=14, bold=True)

cell_4 = ws.cell(row=3, column=1)
cell_4.value = '方位区分:东西'
ws.merge_cells(start_row=3, end_row=3, start_column=1, end_column=3)
cell_4.alignment = Alignment(horizontal='left', vertical='center')
cell_4.font = Font('黑体', size=14, bold=True)

cell_5 = ws.cell(row=3, column=4)
cell_5.value = '轴线明细:无'
ws.merge_cells(start_row=3, end_row=3, start_column=4, end_column=5)
cell_5.alignment = Alignment(horizontal='left', vertical='center')
cell_5.font = Font('黑体', size=14, bold=True)

cell_6 = ws.cell(row=3, column=6)
cell_6.value = '区域明细:1-2'
cell_6.alignment = Alignment(horizontal='left', vertical='center')
cell_6.font = Font('黑体', size=14, bold=True)

cell_7 = ws.cell(row=4, column=1)
cell_7.value = '一级序号:无'
ws.merge_cells(start_row=4, end_row=4, start_column=1, end_column=3)
cell_7.alignment = Alignment(horizontal='left', vertical='center')
cell_7.font = Font('黑体', size=14, bold=True)

cell_8 = ws.cell(row=4, column=4)
cell_8.value = '二级序号:无'
ws.merge_cells(start_row=4, end_row=4, start_column=4, end_column=5)
cell_8.alignment = Alignment(horizontal='left', vertical='center')
cell_8.font = Font('黑体', size=14, bold=True)

__ = ['零件代号', '零件规格', '单位重量', '计重属性名称', '零件明细备注', '零件缩略图']
for i, key in enumerate(__):
    cell_9 = ws.cell(row=5, column=i + 1)
    cell_9.value = key
    cell_9.alignment = Alignment(horizontal='center', vertical='center')
    cell_9.font = Font('黑体', size=11, bold=True)
    cell_9.border = border_style_1

img = Image('1.jpeg')
img.width = 160
img.height = 110
data = ['JA', '50*3', '2.33', '限制(kg)', '角钢(50*3)', img]

for i, key in enumerate(data):
    for row in range(6, 10, 1):
        cell_10 = ws.cell(row=row, column=i + 1)
        if isinstance(key, Image):
            # ws.add_image(key, 'F' + str(row))
            ...
        else:
            cell_10.value = key
            cell_10.font = Font('黑体', size=10, bold=False)
        cell_10.alignment = Alignment(horizontal='center', vertical='center')
        cell_10.border = border_style_2

cell_11 = ws.cell(row=10, column=1)
cell_11.value = '零件种类:4'
cell_11.font = Font('黑体', size=11, bold=True)
cell_11.alignment = Alignment(horizontal='center', vertical='center')

height = ['33', '25', '28', '84', '25']

for row in range(1, 11, 1):
    if row == 1:
        ws.row_dimensions[row].height = height[0]
    elif row <= 4:
        ws.row_dimensions[row].height = height[1]
    elif row == 5:
        ws.row_dimensions[row].height = height[2]
    elif row <= 9:
        ws.row_dimensions[row].height = height[3]
    elif row == 10:
        ws.row_dimensions[row].height = height[4]
ws.column_dimensions['A'].width = 11.625
ws.column_dimensions['B'].width = 11.625
ws.column_dimensions['C'].width = 11.625
ws.column_dimensions['D'].width = 11.625
ws.column_dimensions['E'].width = 17
ws.column_dimensions['F'].width = 24

img = Image('1.jpeg')
img.width = 160
img.height = 110
img2 = copy.deepcopy(img)
img3 = copy.deepcopy(img)
img4 = copy.deepcopy(img)
ws.add_image(img, 'F6')
ws.add_image(img2, 'F7')
ws.add_image(img3, 'F8')
ws.add_image(img4, 'F9')

wb.save('t1.xlsx')


