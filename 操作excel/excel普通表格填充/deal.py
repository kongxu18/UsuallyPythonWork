import pymssql
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils import get_column_letter, column_index_from_string

pd.set_option('display.max_columns', 30)
pd.set_option('display.max_rows', 50)

border_style = Border(left=Side(border_style='thin', color='000000'),
                      right=Side(border_style='thin', color='000000'),
                      top=Side(border_style='thin', color='000000'),
                      bottom=Side(border_style='thin', color='000000'))

alignment_center = Alignment(horizontal='center', vertical='center', wrapText=True)
alignment_left = Alignment(horizontal='left', vertical='center', wrapText=True)

fill_red = PatternFill(fill_type="solid", fgColor="EE292B")
fill_green = PatternFill(fill_type="solid", fgColor="7FCA40")
fill_yellow = PatternFill(fill_type="solid", fgColor='EAC121')
fill_gray = PatternFill(fill_type="solid", fgColor='A6A6A6')
fill_blue = PatternFill(fill_type="solid", fgColor='8DB4E2')


def connection(sql):
    # 创建连接对象
    conn = pymssql.connect(host='erp.highbird.cn', server='erp.highbird.cn', port='9155', user='nizihua',
                           password='13o84x',
                           database='base1')

    cursor = conn.cursor(as_dict=True)

    cursor.execute(sql)
    # 列名
    columns = [name[0] for name in cursor.description]
    df = pd.DataFrame(cursor.fetchall(), columns=columns)
    length = df.shape[0]
    print(length)
    # print(df)
    return df


def set_style(cell, size=12):
    cell.font = Font(name='宋体', bold=True, size=size)
    cell.border = border_style
    # cell.quotePrefix = False
    cell.alignment = alignment_center
    # cell.fill = fill_blue


df_f1 = pd.read_excel('40拉杆采购清单表.xlsx', sheet_name='F1')
df_f2 = pd.read_excel('40拉杆采购清单表.xlsx', sheet_name='F2')
df_f3 = pd.read_excel('40拉杆采购清单表.xlsx', sheet_name='F3')

print(df_f1)
ROWS = set()
COLS = set()


def excel(df: pd.DataFrame, form):
    # df = df.sort_values(by=['行号'],axis=0,ascending=False)
    wb = Workbook()
    sheet = wb.active

    diff_min_col = df.loc[:, '部位代码'].min() - 1
    # diff_min_row = df.loc[:, '一级序号'].min()
    max_row = df.loc[:, '一级序号'].max() + 1
    col_index = 0
    for index, row in df.iterrows():
        col_num = row['部位代码']

        row_num = row['一级序号']
        val = row['构件完整编号']
        if form == 1:
            df_D = df_f1
        elif form == 2:
            df_D = df_f2
        else:
            df_D = df_f3
        design_len = df_D[df_D.iloc[:, 1] == val].iloc[:,2].values[0]
        print(type(design_len),design_len)


        if col_num not in COLS:
            col_index += 1
            # 列 行
            cell_col = sheet.cell(row=1, column=col_index + 1, value='A' + str(col_num))
            letter = get_column_letter(col_index + 1)
            sheet.column_dimensions[letter].width = 27
            sheet.column_dimensions['A'].width = 14
            sheet.row_dimensions[1].height = 30
            set_style(cell_col, 13)
            COLS.add(col_num)

        cell = sheet.cell(row=max_row + 1 - row_num, column=col_index + 1)
        if not cell.value:
            cell.value = val+'/'+str(design_len+138)+'('+str(design_len)+')'
            set_style(cell)

        if row_num not in ROWS:
            cell_row = sheet.cell(row=max_row + 1 - row_num, column=1, value=row_num)
            sheet.row_dimensions[max_row + 1 - row_num].height = 25
            set_style(cell_row, 13)
            ROWS.add(row_num)

    wb.save('t3.xlsx')


if __name__ == '__main__':
    # sql = "select t1.全局分类名称 as 层,t1.一级序号 as 列,二级序号 as 行号,t1.材料名称 as 框架种类,构件完整编号 as 构件编号 " \
    #       "from FT254C构件上传记录(10429,'*','*') as t1 where 材料名称 like 'K%' and t1.全局代码=1 order by 一级序号,二级序号 desc"
    sql = "select 构件完整编号,t1.部位代码,一级序号 from FT254C构件上传记录(10429,53,'*') as t1 " \
          "inner join T257C大连索网 on T257C大连索网.全局代码 = t1.全局代码 and [轴号数值] = t1.部位代码 " \
          "where t1.全局代码 = 3 order by T257C大连索网.部位代码,一级序号"
    df = connection(sql)
    # print(df)
    excel(df,3)
    # print(df_f1)
