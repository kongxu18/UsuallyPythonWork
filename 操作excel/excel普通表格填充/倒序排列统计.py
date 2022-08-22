import pymssql
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils import get_column_letter, column_index_from_string
import copy

pd.set_option('display.max_columns', 1000)  # 显示完整的列
pd.set_option('display.max_rows', None)  # 显示完整的行
pd.set_option('display.width', 1000)  # 显示最大的行宽
pd.set_option('display.max_colwidth', 1000)  # 显示最大的列宽

border_style = Border(left=Side(border_style='thin', color='000000'),
                      right=Side(border_style='thin', color='000000'),
                      top=Side(border_style='thin', color='000000'),
                      bottom=Side(border_style='thin', color='000000'))

border_style2 = Border(left=Side(border_style='thin', color='000000'),
                       right=Side(border_style='thin', color='000000'),
                       top=Side(border_style='thin', color='000000'),
                       bottom=Side(border_style='thin', color='000000'),
                       diagonal=Side(border_style='thin', color='000000'),
                       diagonalUp=False, diagonalDown=True)

alignment_center = Alignment(horizontal='center', vertical='center', wrapText=True)
alignment_left = Alignment(horizontal='left', vertical='center', wrapText=True)
alignment_justify = Alignment(horizontal='justify', vertical='justify', wrap_text=True)
fill_red = PatternFill(fill_type="solid", fgColor="EE292B")
fill_green = PatternFill(fill_type="solid", fgColor="7FCA40")
fill_yellow = PatternFill(fill_type="solid", fgColor='EAC121')
fill_gray = PatternFill(fill_type="solid", fgColor='A6A6A6')
fill_blue = PatternFill(fill_type="solid", fgColor='8DB4E2')


def connection(sql):
    # 创建连接对象
    conn = pymssql.connect(host='erp.highbird.cn', server='erp.highbird.cn', port='9155', user='nizihua',
                           password='13o84x',
                           database='base1', charset='utf8')

    cursor = conn.cursor(as_dict=True)

    cursor.execute(sql)
    # 列名
    columns = [name[0] for name in cursor.description]
    df = pd.DataFrame(cursor.fetchall(), columns=columns)
    length = df.shape[0]
    return df


def statistics(df: pd.DataFrame):
    """
    统计构建状态,用于添加最后部分的统计状态
    """
    group = df.groupby(['一级序号', '构件生存状态'])['构件生存状态'].count().reset_index(name='count')
    return group


def group_by(df: pd.DataFrame):
    group = df.groupby(['构件生存状态'])['构件生存状态'].count().reset_index(name='sum')
    return group


def set_style(cell, size=12):
    cell.font = Font(name='宋体', bold=False, size=size)
    cell.border = border_style
    # cell.quotePrefix = False
    cell.alignment = alignment_center
    # cell.fill = fill_blue


def A1(sheet):
    cell = sheet.cell(row=1, column=1)
    cell.value = '        柱行    '
    cell.border = border_style2
    cell.font = Font(name='宋体', bold=False, size=14)
    cell.alignment = alignment_justify


def whole_style(sheet, w, h):
    excel_w = max(w)
    excel_h = max(h)

    for w in range(1, excel_w + 1):
        for h in range(1, excel_h + 1):
            if w == 1 and h == 1:
                ...
            else:
                cell = sheet.cell(row=h, column=w)
                cell.border = border_style


def excel(df: pd.DataFrame, save_name):
    ROWS = set()
    COLS = set()
    index_dict: dict = {}
    # df = df.sort_values(by=['行号'],axis=0,ascending=False)
    wb = Workbook()
    sheet = wb.active
    sheet.column_dimensions['A'].width = 14
    sheet.row_dimensions[1].height = 37

    # 设置a1 的样式
    A1(sheet)

    diff_min_col = df.loc[:, '一级序号'].min()
    # 最大行数
    max_row = df.loc[:, '二级序号'].max()
    print(max_row, 'row')
    # 醉倒列数
    max_col = df.loc[:, '一级序号'].max()

    # 统计数据
    data: pd.DataFrame = group_by(df)

    data = data.append({'构件生存状态': '构建模型', 'sum': 3}, ignore_index=True)
    print(data)

    # 主题内容添加
    for _, row in df.iterrows():
        # 行列对应的坐标
        col_num = row['一级序号']
        row_num = row['二级序号']
        val = row['构件全称缓存']
        state = row['构件生存状态']

        # 第一行的列名
        # 定位到excel 中的列
        col_index = col_num + 2 - diff_min_col
        # 定位到excel 中的行,需要颠倒排序
        row_index = max_row + 2 - row_num

        # 主体内容设置
        cell = sheet.cell(row=row_index, column=col_index)
        if not cell.value:
            cell.value = val + '            ' + state
            set_style(cell)

        if col_index not in COLS:
            column_name = sheet.cell(row=1, column=col_index, value='A' + str(col_num))
            letter = get_column_letter(col_index)
            sheet.column_dimensions[letter].width = 27
            set_style(column_name, 13)
            COLS.add(col_index)

        if row_index not in ROWS:
            cell_row = sheet.cell(row=row_index, column=1, value=row_num)
            sheet.row_dimensions[row_index].height = 28
            set_style(cell_row, 13)
            ROWS.add(row_index)

        # 获取当前对应的统计生存状态结果
        # res_df = data[data['一级序号'] == col_num]
        # for i, r in res_df.iterrows():
        #     # 要在最下方添加统计
        #     name = r['构件生存状态']
        #     count = r['count']
        #     if name not in index_dict:
        #         index_dict[name] = len(index_dict) + max_row + 2
        #
        #     row_i = index_dict.get(name)
        #     ROWS.add(row_i)
        #
        #     cell = sheet.cell(row=row_i, column=col_index, value=count)
        #     set_style(cell, 13)

    # 单独再第一列添加
    for i, r in data.iterrows():
        name = r['构件生存状态']
        sum = r['sum']

        t_index = int(str(i)) + 2 + max_row
        ROWS.add(t_index)

        cell_1 = sheet.cell(row=t_index, column=1, value=name+'总计')
        sheet.row_dimensions[t_index].height = 25
        set_style(cell_1)

        cell_2 = sheet.cell(row=t_index, column=2, value=sum)
        sheet.row_dimensions[t_index].height = 25
        set_style(cell_2)

    print(ROWS)
    whole_style(sheet, COLS, ROWS)

    wb.save('%s.xlsx' % save_name)


if __name__ == '__main__':
    arg = 'K', 1
    sql = "select * from [FT254E按材料状态速查]('%s',%d)where 构件生存状态代码 > 1 order by 全局代码,部位代码,方位名称,一级序号,二级序号" % arg

    df = connection(sql)
    # print(df)
    # print(df.loc[:,'构件全称缓存'])
    excel(df, 'test')
