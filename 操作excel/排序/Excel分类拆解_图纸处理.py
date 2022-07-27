import json
import sys
import io
import requests
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
import json
import sys
import os
import re
import functools
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import time

日期标识 = time.strftime('%Y%m%d')


def GenBorder(left_style, right_style, top_style, bottom_style):
    rltBorder = Border()
    if left_style != None:
        rltBorder.left = Side(border_style=left_style, color='000000')
    if right_style != None:
        rltBorder.right = Side(border_style=right_style, color='000000')
    if top_style != None:
        rltBorder.top = Side(border_style=top_style, color='000000')
    if bottom_style != None:
        rltBorder.bottom = Side(border_style=bottom_style, color='000000')
    return rltBorder


titleFill = PatternFill(fill_type='solid', start_color='CCCCCC', end_color='000000')
oddFill = PatternFill(fill_type='solid', start_color='EEEEEE', end_color='000000')

headerUseFont = Font(name='宋体', size=20)
contentUseFont = Font(name='宋体', size=11)
titleUseFont = Font(name='宋体', size=12, bold=True)


def initHeaderCell(cell):
    cell.border = GenBorder('thin', 'thin', 'thin', 'thin')
    cell.font = headerUseFont


def initContentCell(cell):
    cell.border = GenBorder('thin', 'thin', 'thin', 'thin')
    cell.font = contentUseFont
    if (cell.row % 2 == 0):
        cell.fill = oddFill


def initTitleCell(cell):
    cell.border = GenBorder('thin', 'thin', 'thin', 'thin')
    cell.fill = titleFill
    cell.font = titleUseFont
    cell.alignment = Alignment(vertical="center", horizontal="center")


class C单元:
    def __init__(self, 序号, data_arr):
        self.序号 = int(序号)
        self.data_arr = data_arr


def sort序号(a, b):
    return a.序号 - b.序号


all单元_arr = []


def 分拆预应力(srcFilePath):
    fileName = os.path.splitext(os.path.basename(srcFilePath))[0]
    print('process:%s' % fileName)
    # print(addQuet)
    srcWB = load_workbook(srcFilePath)
    srcSheet = srcWB['Sheet1']
    wb = Workbook()
    useSheet = wb.active
    columns_arr = []
    for srcSheetName in srcWB.sheetnames:
        srcSheet = srcWB[srcSheetName]
        maxRow = srcSheet.max_row
        maxColumn = srcSheet.max_column

        for row in range(1, maxRow + 1):
            print('%d/%d' % (row, maxRow))
            for col in range(0, maxColumn, 5):
                cell_0 = srcSheet[row][col]
                cell_1 = srcSheet[row][col + 1]
                cell_2 = srcSheet[row][col + 2]
                cell_3 = srcSheet[row][col + 3]
                cell_4 = srcSheet[row][col + 4]
                if cell_0.value == '单元编号':
                    if len(columns_arr) == 0:
                        columns_arr = [cell_0.value, cell_1.value, cell_2.value, cell_3.value, cell_4.value]
                    continue

                all单元_arr.append(
                    C单元(cell_0.value, [cell_0.value, cell_1.value, cell_2.value, cell_3.value, cell_4.value]))
            # if row > 2:
            #     break
    all单元_arr.sort(key=functools.cmp_to_key(sort序号))
    for i in range(len(columns_arr)):
        initTitleCell(useSheet.cell(row=1, column=1 + i, value=columns_arr[i]))
    for 单元i, 单元 in enumerate(all单元_arr):
        for i in range(len(单元.data_arr)):
            initContentCell(useSheet.cell(row=2 + 单元i, column=1 + i, value=单元.data_arr[i]))
    wb.save(os.path.join("D:\\rhinoWork\\wantexcel\\已拆解\\", '图纸应力_已排序.xlsx'))


dirPath = 'D:\\rhinoWork\\wantexcel'
分拆预应力(os.path.join(dirPath, '图纸预应力.xlsx'))
