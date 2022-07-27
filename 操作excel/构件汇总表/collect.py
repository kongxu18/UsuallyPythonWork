import pymssql as mssql

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string
import numpy as np
import pandas as pd
import os
from openpyxl.drawing.image import Image

pd.set_option('display.max_columns', 100)
pd.set_option('display.max_rows', 50)
pd.set_option('display.width', 1000)
border_style = Border(left=Side(border_style='thin', color='000000'),
                      right=Side(border_style='thin', color='000000'),
                      top=Side(border_style='thin', color='000000'),
                      bottom=Side(border_style='thin', color='000000'))

alignment_center = Alignment(horizontal='center', vertical='center', wrapText=True)
alignment_left = Alignment(horizontal='left', vertical='center', wrapText=True)
alignment_right = Alignment(horizontal='right', vertical='center', wrapText=True)
alignment_right_bottom = Alignment(horizontal='right', vertical='bottom', wrapText=True)

fill_red = PatternFill(fill_type="solid", fgColor="EE292B")
fill_green = PatternFill(fill_type="solid", fgColor="7FCA40")
fill_yellow = PatternFill(fill_type="solid", fgColor='EAC121')
fill_gray = PatternFill(fill_type="solid", fgColor='A6A6A6')
fill_blue = PatternFill(fill_type="solid", fgColor='8DB4E2')


class Row(object):
    def __init__(self, row):
        self.__row = row

    @property
    def val(self):
        return self.__row

    @val.setter
    def val(self, num):
        self.__row += num


class SqlData(object):
    def __init__(self):
        self.__data = None
        self.__oldDf = None
        self.__color_list = {}

        self.hide_name = []
        self.info_start_row = 3

    def connect_sql(self, **kwargs):

        project = kwargs.get('project')

        try:
            # sql服务器名，这里(127.0.0.1)是本地数据库IP
            serverName = 'erp.highbird.cn:9155'
            # 登陆用户名和密码
            userName = 'nizihua'
            passWord = '333333'
            # 建立连接并获取cursor
            conn = mssql.connect(serverName, userName, passWord, "base1", 'utf8')
            cursor = conn.cursor()

            sql = "select  * from FT254C构件数据统计(%d) order by 完整编号 ,参数详称"
            cursor.execute(sql, project)
            # 列名
            columns = [name[0] for name in cursor.description]

            df = pd.DataFrame(cursor.fetchall(), columns=columns)
            self.__oldDf = df.copy(deep=True)
            # self.hide_col(df)
            cursor.close()
            conn.close()
        except Exception as err:
            print(err, 'sql---错误')

    @property
    def col_count(self):
        return 7

    @property
    def group_data(self):
        df = self.__oldDf
        if df.empty:
            return {}
        group = df.groupby('完整编号')
        return group, Row(self.info_start_row)

    # def info_row(self):
    #     row, counter = self.info_start_row, 0
    #     while True:
    #         yield row
    #         row += counter


class Header(object):
    BASICS = '大连梭鱼湾专业足球场立面索膜结构——加工施工质量记录'

    def __init__(self, *args, **kwargs):
        self.Excel = kwargs.get('cls')
        self.sql_data = kwargs.get('sql_data')
        self.set_params()
        self.start_row = 1
        self.end_column = 7

    def set_params(self):
        self.sheet = self.Excel.sheet

    def set_style(self, cell):
        cell.font = Font(name='宋体', bold=True, size=18)
        # cell.quotePrefix = False
        cell.alignment = alignment_right

    def set_height(self):
        self.sheet.row_dimensions[1].height = 50

    def add(self, **kwargs):
        if not self.sheet:
            return
        cell = self.sheet.cell(row=self.start_row, column=1, value=self.BASICS)
        # 合并单元格
        self.sheet.merge_cells(start_row=self.start_row, start_column=1, end_row=self.start_row,
                               end_column=self.end_column)
        self.set_style(cell)


class Page(object):

    def __init__(self, *args, **kwargs):
        self.Excel = kwargs.get('cls')
        self.sql_data = kwargs.get('sql_data')
        self.set_params()

    def set_params(self):
        self.sheet = self.Excel.sheet
        self.data = self.sql_data.group_data

    def add(self):
        if self.data:
            # 遍历数据增加 info table
            ...
            group_data, row = self.data
            for component_name, components_frame in group_data:
                info = Info(excel=self.Excel, words=component_name, row=row)
                info.add()

                table = Table(excel=self.Excel, table=components_frame, row=row)
                table.add()


class Info(object):
    BASICS = '构件编号:'

    def __init__(self, *args, **kwargs):
        self.Excel = kwargs.get('excel')
        self.row_object = kwargs.get('row')
        self.set_params()

        self.words = kwargs.get('words')
        self.end_column = 7

    def set_params(self):
        self.sheet = self.Excel.sheet
        self.row = self.row_object.val - 1

    def set_style(self, cell):
        cell.font = Font(name='宋体', bold=True, size=11)
        # cell.quotePrefix = False
        cell.alignment = alignment_right_bottom

    def set_height(self):

        self.sheet.row_dimensions[self.row].height = 34

    def add(self):
        if not self.sheet:
            return
        if self.row and self.words:
            cell = self.sheet.cell(row=self.row, column=1, value=self.BASICS + self.words)
            # 合并单元格
            self.sheet.merge_cells(start_row=self.row, start_column=1, end_row=self.row,
                                   end_column=self.end_column)
            self.set_style(cell)
            self.set_height()


class Table(object):
    def __init__(self, *args, **kwargs):
        self.Excel = kwargs.get('excel')
        self.row_object = kwargs.get('row')
        self.table = kwargs.get('table')
        self.set_params()
        self.end_column = 7
        self.constituent_part = []

    def set_params(self):
        self.sheet = self.Excel.sheet
        self.row = self.row_object.val

    def set_height(self, row):
        # self.sheet.row_dimensions[row - 1].height = 34
        self.sheet.row_dimensions[row].height = 20
        self.sheet.row_dimensions[row + 1].height = 145
        self.sheet.row_dimensions[row + 2].height = 20
        self.sheet.row_dimensions[row + 3].height = 20

    def set_style(self, cell):
        cell.font = Font(name='宋体', bold=True, size=11)
        # cell.quotePrefix = False
        cell.border = border_style

        cell.alignment = alignment_center

    def data(self):
        df = self.table
        line_style = df.loc[df.参数详称 == '详图序号']
        张力 = df.loc[df.参数详称 == 'F']
        总长 = df.loc[df.参数详称 == 'L']
        B端 = df.loc[df.参数详称 == 'B端']
        T端 = df.loc[df.参数详称 == 'T端']
        c_df = df.query("参数详称!= '详图序号' and 参数详称 != 'F' and 参数详称 != 'L' and 参数详称 != 'B端' and 参数详称 != 'T端'")
        c_df = c_df.sort_values(by='参数序号')
        c_length = c_df.shape[0]

        if line_style.empty:
            line_style = pd.DataFrame({'完整编号': [''], '参数详称': ['详图序号'], '参数值': [2], '复核偏差': [-1], '文件路径': ['']})

        row = pd.concat([line_style, 张力, 总长, B端, c_df, T端], ignore_index=True)

        # 6个数据为一组
        res_array = []
        row_index = self.row
        start, counter = 0, 5
        while start < row.shape[0]:
            res_array.append((row_index, row.loc[start:start + counter]))
            start += counter + 1

        return res_array

    def add(self):
        array = self.data()
        for row_index, data in array:

            if data.empty:
                empty = True
            else:
                empty = False
            self.constituent_part.append(TableHeader(excel=self.Excel, row=self.row_object, empty=empty))
            self.constituent_part.append(TableBody(excel=self.Excel, row=self.row_object, data=data))
            self.row_object.val = 5
        for part in self.constituent_part:
            part.add()


class TableHeader(Table):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.empty = kwargs.get('empty')

    def add(self):
        row = self.row
        params = ['参数', '测量照片', '设计值', '偏差值']
        for i, word in enumerate(params):
            if not self.empty:
                cell = self.sheet.cell(row=row + i, column=1, value=word)
                self.set_style(cell)
        self.set_height(row)


class TableBody(Table):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.df = kwargs.get('data')

    def add(self):
        row = self.row
        col = 2

        for index, val in self.df.iterrows():
            参数 = val['参数详称']
            if 参数 == '详图序号':
                参数 = '索样式'
            elif 参数 == 'F':
                参数 = '张力'
            elif 参数 == 'L':
                参数 = '总长'
            照片 = str(val['文件路径'])

            设计值 = str(val['参数值']) if int(val['参数值']) >= 0 else ''

            偏差值 = str(val['复核偏差']) if int(val['复核偏差']) >= 0 else ''

            params = [参数, 照片, 设计值, 偏差值]
            img_arr = ['钢索图片/钢索详图1.jpg', '钢索图片/钢索详图2.jpg', '钢索图片/钢索详图3.jpg', '钢索图片/钢索详图4.jpg']
            for i, word in enumerate(params):
                if i == 1:
                    # 图片
                    if 设计值 == '' or 设计值 == '2':
                        path = img_arr[1]
                    else:
                        path = img_arr[int(设计值)]
                    img = Image(path)
                    img.width = 100
                    img.height = 100
                    # self.sheet.
                cell = self.sheet.cell(row=row + i, column=col, value=word)
                self.set_style(cell)
            col += 1


class Excel(object):
    def __init__(self):
        self.components = []
        self.__params = None
        self.__wb, self.__sheet = self.create_workbook()

    @property
    def params(self, **kwargs):
        return self.__params

    @params.setter
    def params(self, params: dict = None):
        self.__params = params

    @staticmethod
    def create_workbook(parma=''):
        wb = load_workbook('sample' + parma + '.xlsx')
        sheet = wb.active
        return wb, sheet

    @property
    def sheet(self):
        return self.__sheet

    def set_col_width(self, sql_data):

        # width_list = sql_data.max_width
        # for i, val in enumerate(width_list):
        #     letter = get_column_letter(i + 1)
        #     self.sheet.column_dimensions[letter].width = val
        ...

    def create_excel(self, path, sql_data=None):
        if self.components:

            for component in self.components:
                handler = component(cls=self, sql_data=sql_data)
                handler.add()
        else:
            return

        self.set_col_width(sql_data)
        self.__wb.save(path)


if __name__ == '__main__':
    s = SqlData()
    s.hide_name = []
    s.connect_sql(project=10429)

    e = Excel()
    e.components = [Header, Page]
    # e.params = {'title': '天地无限宇宙', 'project': '浩大无敌工程', 'processingData': '2088-03-03'}
    e.create_excel('a.xlsx', s)
