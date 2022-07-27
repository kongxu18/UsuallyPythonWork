"""
连接sql
"""
# -*- coding: UTF-8 –*-
import pymssql as mssql
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string
import numpy as np
from PIL import ImageGrab

# import xlwings
# 根据列的数字返回字母
# print(get_column_letter(2))
# 根据字母返回列的数字
# print(column_index_from_string('D'))
pd.set_option('display.max_columns', 30)
pd.set_option('display.max_rows', 50)

border_style = Border(left=Side(border_style='thin', color='000000'),
                      right=Side(border_style='thin', color='000000'),
                      top=Side(border_style='thin', color='000000'),
                      bottom=Side(border_style='thin', color='000000'))

alignment_center = Alignment(horizontal='center', vertical='center', wrapText=True)
alignment_left = Alignment(horizontal='left', vertical='center', wrapText=True)

fill_red = PatternFill(fill_type="solid", fgColor="EE292B")
fill_green = PatternFill(fill_type="solid", fgColor="7FCA40")
fill_yellow = PatternFill(fill_type="solid", fgColor='EAC121')
fill_gray = PatternFill(fill_type="solid", fgColor='A6A6A6')
fill_blue = PatternFill(fill_type="solid", fgColor='8DB4E2')


class SqlData(object):
    def __init__(self):
        self.__data = None
        self.__oldDf = None
        self.__color_list = {}

        self.hide_name = []

    @property
    def data(self):
        return self.__data

    def hide_col(self, df):
        for name in self.hide_name:
            df.drop(name, axis=1, inplace=True)
        self.__data = df

    def connect_sql(self, **kwargs):
        title = kwargs.get('title')
        project = kwargs.get('project')
        processing_data = kwargs.get('data')

        try:
            # sql服务器名，这里(127.0.0.1)是本地数据库IP
            serverName = '192.168.0.202:9155'
            # 登陆用户名和密码
            userName = 'shilei'
            passWord = 'Sl941521'
            # 建立连接并获取cursor
            conn = mssql.connect(serverName, userName, passWord, "base1", 'utf8')
            cursor = conn.cursor()
            sql = 'EXEC P255E构件工序加工进度 %d,%d,%s;'
            cursor.execute(sql, (title, project, processing_data))

            # 列名

            columns = [name[0] for name in cursor.description]

            df = pd.DataFrame(cursor.fetchall(), columns=columns)
            self.__oldDf = df.copy(deep=True)
            self.hide_col(df)
            cursor.close()
            conn.close()
        except Exception as err:
            print(err, 'sql---错误')

    @staticmethod
    def isChinese(word):
        count = 0
        if not word:
            return 0

        word = str(word)

        try:
            for ch in word:
                if '\u4e00' <= ch <= '\u9fff':
                    count += 2 * 1.5
                else:
                    count += 1 * 1.5
        except Exception as err:
            print(word, '--ischinese')
        return count

    @property
    def max_width(self):
        thead_list = []
        tbody_list = []
        df = self.data
        for i in range(0, df.shape[1]):
            # 标题的宽度
            thead = df.columns[i]

            thead_width = len(thead) * 2 * 1.75 + 1
            thead_list.append(thead_width)
            # 内容的宽度

            df_tamp = df.iloc[:, i]
            tbody_width = df_tamp.map(lambda x: self.isChinese(x) + 1).max()
            tbody_list.append(tbody_width)

        res_list = np.maximum(thead_list, tbody_list)
        return res_list

    @property
    def apply_color_list(self):
        df = self.data
        cursor = None

        for i in range(0, df.shape[1]):
            empty = np.empty(shape=df.shape[0], dtype='U10')
            thead = df.columns[i]
            if thead == '加工工序大类':
                cursor = i
            if cursor and cursor < i:
                df_tamp = df.iloc[:, i]
                tamp_list = self.__color_list.get(thead) if self.__color_list.get(thead) else empty
                # excel 索引从0开始，这里数据输出+1

                self.greater_than_zero(df_tamp, tamp_list, i + 1)
                self.equal_one(df_tamp, tamp_list, i + 1)
                self.equal_zero(df_tamp, tamp_list, i + 1)

        return self.__color_list

    def greater_than_zero(self, series, tamp_list, key):
        mask = (series > 0) & (series < 1)
        tamp_list[mask] = 'yellow'

        self.__color_list.update({key: list(tamp_list)})

    def equal_one(self, series, tamp_list, key):
        mask = series == 1
        tamp_list[mask] = 'green'
        self.__color_list.update({key: list(tamp_list)})

    def equal_zero(self, series, tamp_list, key):
        mask = series == 0
        tamp_list[mask] = 'red'
        self.__color_list.update({key: list(tamp_list)})

    @property
    def col_count(self):
        df = self.data
        count = df.shape[1]
        return count

    @property
    def params(self):
        df = self.__oldDf
        if df.empty:
            return {}
        col_name = ['项目登记名称', '项目单体名称', '加工日期']
        title = df['项目登记名称'][0]
        project = df['项目单体名称'][0]
        data = df['加工日期'][0]
        return {'title': title, 'project': project, 'data': data}


class Header(object):
    BASICS = '项目钢材生产进度表'

    def __init__(self, *args, **kwargs):
        self.Excel = kwargs.get('cls')
        self.sql_data = kwargs.get('sql_data')
        self.wb, self.sheet = None, None
        self.title = None
        self.col_count = None
        self.set_params()
        self.row = 1

    def set_params(self):
        self.sheet = self.Excel.sheet
        self.title = self.sql_data.params.get('title')
        self.col_count = self.sql_data.col_count

    @property
    def full_name(self):
        return self.title if self.title else '公共' + self.BASICS

    def set_style(self, cell):
        cell.font = Font(name='宋体', bold=True, size=24)
        cell.border = border_style
        # cell.quotePrefix = False
        cell.alignment = alignment_center

    def set_height(self):
        self.sheet.row_dimensions[1].height = 50

    def add(self, **kwargs):
        if not self.sheet:
            return
            # 默认就是第一行定死
        cell = self.sheet.cell(row=self.row, column=1, value=self.full_name)
        if not self.col_count:
            raise TypeError('col_count 列数为0')
        for i in range(self.col_count):
            cell_ = self.sheet.cell(row=self.row, column=1 + i)
            cell_.border = border_style
        # 合并单元格
        self.sheet.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=self.col_count)
        # 设置单元格格式
        self.set_style(cell)
        self.set_height()


class Info(Header):
    BASICS = '项目单体的名称:{project}      进度加工日期:{data}'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.row = 2

    def set_params(self):
        self.sheet = self.Excel.sheet
        self.project = self.sql_data.params.get('project')
        self.date = self.sql_data.params.get('data')
        self.col_count = self.sql_data.col_count

    def set_style(self, cell):
        cell.font = Font(name='宋体', bold=True, size=20)
        cell.border = border_style
        # cell.quotePrefix = False
        cell.alignment = alignment_center
        cell.fill = fill_blue

    def set_height(self):
        self.sheet.row_dimensions[2].height = 50

    @property
    def full_name(self):
        if self.project:
            self.BASICS = self.BASICS.replace('{project}', self.project)
        if self.date:
            self.BASICS = self.BASICS.replace('{data}', self.date)
        return self.BASICS


class Table(object):
    def __init__(self, *args, **kwargs):
        self.Excel = kwargs.get('cls')
        self.sql_data = kwargs.get('sql_data')
        self.wb, self.sheet = None, None
        self.df = None
        self.set_params()
        self.row = 3

    def set_params(self):
        self.sheet = self.Excel.sheet
        self.df = self.sql_data.data

    def set_height(self):
        self.sheet.row_dimensions[3].height = 16

    def set_style(self):
        full_color_list = self.sql_data.apply_color_list
        # print(self.df.shape[0], self.row)
        for col in range(1, self.df.shape[1] + 1):
            # 表头样式
            cell = self.sheet.cell(row=self.row, column=col)
            cell.border = border_style
            cell.font = Font(name='宋体', bold=True, size=18)
            cell.alignment = alignment_center

            color_list = full_color_list.get(col)
            for row in range(self.row + 1, self.df.shape[0] + self.row + 1):
                # 表体
                cell = self.sheet.cell(row=row, column=col)
                # 先设置style 不然会覆盖
                if color_list:
                    cell.style = 'Percent'
                    color = color_list[row - self.row - 1]
                    if color == 'yellow':
                        cell.fill = fill_yellow
                    elif color == 'green':
                        cell.fill = fill_green
                    elif color == 'red':
                        cell.fill = fill_red
                    elif color == '':
                        cell.fill = fill_gray

                cell.border = border_style
                cell.font = Font(name='宋体', bold=False, size=16)
                cell.alignment = alignment_center

    def add(self):
        for row in dataframe_to_rows(self.df, index=False, header=True):
            self.sheet.append(row)
        self.set_style()


class Excel(object):
    def __init__(self):
        self.components = []
        self.__params = None
        self.__wb, self.__sheet = self.create_workbook()

    @property
    def params(self, **kwargs):
        return self.__params

    @params.setter
    def params(self, params: dict = None):
        self.__params = params

    @staticmethod
    def create_workbook():
        wb = Workbook()
        sheet = wb.active
        return wb, sheet

    @property
    def sheet(self):
        return self.__sheet

    def set_col_width(self, sql_data):

        width_list = sql_data.max_width
        for i, val in enumerate(width_list):
            letter = get_column_letter(i + 1)
            self.sheet.column_dimensions[letter].width = val

    def show_img(self):
        self.sheet.Range('A1', 'D5').CopyPicture()

    def create_excel(self, path, sql_data):
        if self.components:

            for component in self.components:
                handler = component(cls=self, sql_data=sql_data)
                handler.add()
        else:
            return

        self.set_col_width(sql_data)
        self.__wb.save(path)


if __name__ == '__main__':
    s = SqlData()
    s.hide_name = ['项目登记名称', '项目单体名称', '加工日期', '项目登记名称代码', '项目构件定义代码']
    s.connect_sql(title=10456, project=6, data='2021-01-01')

    print(s.params)
    # print(s.apply_color_list)
    e = Excel()
    e.components = [Header, Info, Table]
    # e.params = {'title': '天地无限宇宙', 'project': '浩大无敌工程', 'processingData': '2088-03-03'}
    e.create_excel('a.xlsx', s)
