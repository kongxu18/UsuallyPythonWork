from openpyxl.styles import Font, Alignment, Side, Border, PatternFill, GradientFill
from openpyxl import load_workbook


def do(path, name):
    # 只能打开已经存在的表格，不能用该方法创建一个新的表格
    workbook = load_workbook(filename=path)
    for sheetname in workbook.sheetnames:
        sheets = workbook[sheetname]
        sheets['A1'] = name + '水沟，水沟角钩，水沟T铁'

    workbook.save(path)


do('G4_N.xlsx', 'G4北侧')
do('G4_S.xlsx', 'G4南侧')

do('G3_N.xlsx', 'G3北侧')
do('G3_S.xlsx', 'G3南侧')

do('G2_N.xlsx', 'G2北侧')
do('G2_S.xlsx', 'G2南侧')
