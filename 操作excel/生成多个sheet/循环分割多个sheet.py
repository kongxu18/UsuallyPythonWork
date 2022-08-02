import openpyxl
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image
from openpyxl_image_loader import SheetImageLoader

pd.set_option('display.max_columns', 30)
pd.set_option('display.max_rows', 50)

border_style = Border(left=Side(border_style='thin', color='000000'),
                      right=Side(border_style='thin', color='000000'),
                      top=Side(border_style='thin', color='000000'),
                      bottom=Side(border_style='thin', color='000000'))

alignment_center = Alignment(horizontal='center', vertical='center', wrapText=True)
alignment_left = Alignment(horizontal='left', vertical='center', wrapText=True)

fill_red = PatternFill(fill_type="solid", fgColor="EE292B")
fill_green = PatternFill(fill_type="solid", fgColor="7FCA40")
fill_yellow = PatternFill(fill_type="solid", fgColor='EAC121')
fill_gray = PatternFill(fill_type="solid", fgColor='A6A6A6')
fill_blue = PatternFill(fill_type="solid", fgColor='8DB4E2')


# for col in range(cols):
#     series = df.iloc[:, col]
#     print(series.name, '-----')
#     print(series)
#
# excel = openpyxl.load_workbook('F1模板.xlsx')


class Model(object):
    def __init__(self, data, model):
        self.model_excel = openpyxl.load_workbook(model)
        self.df = data

    @property
    def model_sheet(self):
        return self.model_excel.active

    def get_img(self, new_sheet, png):
        # img = Image.open('f1.png')
        # new_sheet.add_image(img, 'A18')
        img = Image(png)
        # print(img.width, img.height)
        img.width, img.height = img.width * 0.63, img.height * 0.66
        # print(img.ref,img.path)

        new_sheet.add_image(img, 'A22')

    def create_sheet(self, sheet_name, sheet_val, png):
        # 复制原有sheet表，创建新sheet
        new_sheet = self.model_excel.copy_worksheet(self.model_sheet)
        self.get_img(new_sheet, png)
        # 为新复制创建的sheet表重命名为遍历的值

        new_sheet.title = sheet_name

        sheet_val = [sheet_name] + list(sheet_val)
        print(sheet_val)
        new_sheet.column_dimensions['A'].width = 14
        new_sheet.column_dimensions['B'].width = 27
        for row, val in enumerate(sheet_val):
            cell = new_sheet.cell(row=row + 1, column=2, value=val)
            self.set_style(cell)
            new_sheet.row_dimensions[row+1].height = 25

        new_sheet.row_dimensions[1].height = 30


    @staticmethod
    def set_style(cell, size=12):
        cell.font = Font(name='宋体', bold=True, size=size)
        cell.border = border_style
        # cell.quotePrefix = False
        cell.alignment = alignment_center

    def creat_workbook(self, path, png):
        # 安装data 列 进行遍历并创建sheet
        rows, cols = self.df.shape
        for col in range(1, cols):
            series = self.df.iloc[:, col]
            col_name = series.name
            self.create_sheet(sheet_name=col_name, sheet_val=series, png=png)

        self.model_excel.save(path)


# data_F1 = pd.read_excel('40拉杆布置图-单尺寸-配图.xlsx', sheet_name='F1')

# m1 = Model(data_F1, 'F1模板.xlsx')
# m1.creat_workbook('F1.xlsx', 'f1.png')

# data_F1 = pd.read_excel('40拉杆布置图-单尺寸-配图.xlsx', sheet_name='F2')
#
# m1 = Model(data_F1, 'F2模板.xlsx')
# m1.creat_workbook('F2.xlsx', 'f2.png')

data_F3 = pd.read_excel('40拉杆布置图-单尺寸-配图.xlsx', sheet_name='F3')

m1 = Model(data_F3, 'F3模板.xlsx')
m1.creat_workbook('F3.xlsx', 'f3.png')

