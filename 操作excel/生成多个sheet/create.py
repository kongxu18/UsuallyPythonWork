import openpyxl
import pandas
import pandas as pd


class Model(object):
    def __init__(self, data_: object):
        self.model_excel = openpyxl.load_workbook('模版.xlsx')
        self.data = data_.data

    @property
    def model_sheet(self):
        return self.model_excel.active

    def create_sheet(self, sheet_name, sheet_val, bool=True):

        max_row, max_col = self.model_sheet.max_row, self.model_sheet.max_column
        # 复制原有sheet表，创建新sheet
        new_sheet = self.model_excel.copy_worksheet(self.model_sheet)
        # 为新复制创建的sheet表重命名为遍历的值
        if bool:
            new_sheet.title = '+W' + sheet_name[:-1]
        else:
            new_sheet.title = '-W' + sheet_name[:-1]

        for row in range(3, len(sheet_val) + 3):
            new_sheet.cell(row=row, column=7, value=sheet_val[row - 3])

    def creat_workbook(self, path, bool):
        for key, val in self.data.items():
            self.create_sheet(key, val, bool)
        self.model_excel.save(path)


class Data(object):
    def __init__(self, path, sheet_name):
        self.sheets = pd.read_excel(path, sheet_name=None)

        self.sheet_name = sheet_name

    @property
    def df(self):
        df = self.sheets[self.sheet_name]
        return df

    @property
    def data(self):
        res = {}
        for col_name in self.df.columns.values:
            df = self.df[col_name] * 0.6
            df = list(df)
            res[col_name] = df
        res.pop('分区')
        return res


if __name__ == '__main__':
    data_obj = Data('云阳遗迹馆膜结构50年回归期风荷载分区极值荷载-HB.xlsx', '极大值荷载')
    # res = data_obj.series_col()
    # print(res)
    e = Model(data_obj)
    e.creat_workbook('正.xlsx',True)

    data_obj = Data('云阳遗迹馆膜结构50年回归期风荷载分区极值荷载-HB.xlsx', '极小值荷载')
    # res = data_obj.series_col()
    # print(res)
    e = Model(data_obj)
    e.creat_workbook('负.xlsx',False)
