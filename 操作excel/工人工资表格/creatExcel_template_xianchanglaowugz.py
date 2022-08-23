import json
import sys
import io
import os
import shutil
import json
import pymssql
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from 操作excel.excelApi.style import *
from copy import copy
import re


class SqlData(object):
    def __init__(self, sql):
        self.sql = sql
        self.__data = None
        self.__len = None
        self.connection()

    def connection(self):
        # 创建连接对象
        try:
            conn = pymssql.connect(host='erp.highbird.cn', server='erp.highbird.cn', port='9155', user='nizihua',
                                   password='13o84x',
                                   database='base1', charset='utf8')

            cursor = conn.cursor(as_dict=True)

            cursor.execute(self.sql)
            columns = [name[0] for name in cursor.description]
            df = pd.DataFrame(cursor.fetchall(), columns=columns)
            self.__len = df.shape[0]
            self.__data = df

        except Exception as err:
            raise ValueError('err:', 'sql交互出现问题 ', str(err))

    @property
    def data(self):
        return self.__data

    @property
    def length(self):
        return self.__len


class DataDeal(object):
    def __init__(self, df):
        self.df = df
        self.set_df()

    def set_df(self):
        if isinstance(self.df, pd.DataFrame):
            if self.df.empty:
                raise ValueError('传入数据为空')
            else:
                self.column_type_revise()
                self.add_new_column()

    def column_type_revise(self):
        """
        对列的类型进行修正
        """
        self.df['工人工日'] = pd.to_numeric(self.df['工人工日'])
        self.df['劳务费小计'] = pd.to_numeric(self.df['劳务费小计'])

    def add_new_column(self):
        """
        拆分 日期增加 日子列
        """
        self.df['month'] = self.df['考勤日期'].astype('str').str.split('-').str.get(1).astype('int')
        self.df['year'] = self.df['考勤日期'].astype('str').str.split('-').str.get(0).astype('int')
        self.df['day'] = self.df['考勤日期'].astype('str').str.split('-').str.get(2).astype('int')

    def group_by_team(self):
        group: pd.DataFrame = self.df.groupby(['班组名称'])
        return group

    def sum_money_by_team_month(self):
        group: pd.DataFrame = self.df.groupby(['班组名称', 'month'])['劳务费小计'].sum().reset_index(name='sum')
        return group

    def sum_money_by_team(self):
        """
        按照班组统计所有金额
        """
        sum_money: pd.DataFrame = self.df.groupby(['班组名称'])['劳务费小计'].sum().reset_index(name='total')
        return sum_money

    def sum_money(self):
        """
        统计所有金额
        """
        sum_money: float = self.df['劳务费小计'].sum()
        return sum_money


def get_color():
    while True:
        yield fill_浅粉
        yield fill_浅蓝


def set_style(cell):
    cell.font = font10
    cell.alignment = alignment_center
    cell.border = border_style


class Clone(object):
    def __init__(self, model_path):
        self.model_path = model_path
        self._model_sheet = None
        self.start_row = 1
        self.start_col = 1
        self._workbook = None
        self.load()
        self.model_sheetNames = ['one', 'two', 'empty']

    def load(self):
        try:
            wb = load_workbook(self.model_path)
            self._workbook = wb
        except Exception as err:
            self._workbook = None
        return self._workbook

    @property
    def workbook(self):
        return self._workbook

    @property
    def model_sheet(self):
        return self._model_sheet

    @model_sheet.setter
    def model_sheet(self, sheetName):
        try:
            sheet = self.workbook[sheetName]
            self._model_sheet = sheet
        except Exception as err:
            self._model_sheet = None

    @staticmethod
    def cloneCell(srcCell, dstCell):
        if srcCell == dstCell:
            return
        dstCell.value = srcCell.value
        if srcCell.has_style:
            dstCell.font = copy(srcCell.font)
            dstCell.border = copy(srcCell.border)
            dstCell.fill = copy(srcCell.fill)
            dstCell.number_format = copy(srcCell.number_format)
            dstCell.protection = copy(srcCell.protection)
            dstCell.alignment = copy(srcCell.alignment)

    def get_merge_info(self):
        return self.model_sheet.merged_cells

    @staticmethod
    def letter_to_num(p):
        """
        传入 A1 输出 （1，1）
        """
        # pattern = re.compile(r'(\D*)(\d*)')

        letter = re.findall('(\D+)(\d+)', p)[0]
        col = column_index_from_string(letter[0])
        row = int(letter[1])
        return row, col

    def deal_merge_info(self, sheet, offset):
        """
        模板从（1，1）开始
        """
        info = self.get_merge_info()
        row_off = offset[0] - 1
        col_off = offset[1] - 1
        for _ in info:
            m_li = str(_).split(':')
            # 获取点的坐标，对点进行实际的偏移计算
            p1, p2 = self.letter_to_num(m_li[0]), self.letter_to_num(m_li[1])

            p1 = p1[0] + row_off, p1[1] + col_off
            p2 = p2[0] + row_off, p2[1] + col_off
            # print(p1, p2)
            sheet.merge_cells(start_row=p1[0], start_column=p1[1], end_row=p2[0], end_column=p2[1])

    def clone(self, rc_range, dst_sheet):
        """
        rc_range 目标sheet 的 范围
        反推 模版的范围
        从 模板 克隆内容到 目标sheet，定位通过左上角 row，col
        """
        # 对角线定位一个范围
        row_s, col_s = rc_range[0]
        row_e, col_e = rc_range[1]

        # 从第一行开始，
        for row in range(self.start_row, row_e - row_s + 2):
            # 行高
            # print(self.model_sheet.row_dimensions[row].height)
            dst_sheet.row_dimensions[row_s + row - 1].height = copy(self.model_sheet.row_dimensions[row].height)

            for col in range(self.start_col, col_e - col_s + 2):
                src_cell = self.model_sheet.cell(row=row, column=col)
                dst_cell = dst_sheet.cell(row=row_s + row - 1, column=col_s + col - 1)
                self.cloneCell(src_cell, dst_cell)

        self.deal_merge_info(dst_sheet, rc_range[0])

    def del_model_sheets(self):
        for name in self.model_sheetNames:
            self.workbook.remove(self.workbook[name])


def clone_header(c, dst_sheet, loc_row):
    """
    header 表头的范围
    """
    # loc_row 起始的行数 ，默认的列为1
    # 模板长度35 ， 高度4行
    p1 = (loc_row, 1)
    p2 = (loc_row + 3, 35)

    rc_range = [p1, p2]

    # 在克隆前，选择需要克隆的sheet
    c.model_sheet = 'one'
    c.clone(rc_range, dst_sheet)

    return loc_row + 4


def clone_footer(c, dst_sheet, loc_row):
    ...
    p1 = (loc_row, 1)
    p2 = (loc_row + 3, 35)

    rc_range = [p1, p2]

    # 在克隆前，选择需要克隆的sheet
    c.model_sheet = 'two'
    c.clone(rc_range, dst_sheet)

    return loc_row + 4


def create_table(projectName, month_word, loc_row, dataframe, cloneObj, dst_sheet):
    loc_row = clone_header(cloneObj, dst_sheet, loc_row)
    cell_project = dst_sheet.cell(row=loc_row - 3, column=2, value=projectName)
    cell_month_word = dst_sheet.cell(row=loc_row - 3, column=29, value=month_word)
    # write_body
    # ----------------------------------
    row_i = loc_row
    group_name = dataframe.groupby('工人姓名')
    for name, df_month in group_name:
        cell_c1 = dst_sheet.cell(row=row_i, column=1, value=name)
        set_style(cell_c1)
        dst_sheet.row_dimensions[row_i].height = 20

        time_sum = (df_month['工人工日'] * 100).sum() / 100
        price = df_month.loc[:, '工日单价'].iloc[0]
        total_price = df_month['劳务费小计'].sum()

        cell_AG = dst_sheet.cell(row=row_i, column=33, value=time_sum)
        set_style(cell_AG)

        cell_AH = dst_sheet.cell(row=row_i, column=34, value=price)
        set_style(cell_AH)

        cell_AI = dst_sheet.cell(row=row_i, column=35, value=total_price)
        set_style(cell_AI)

        # print(df_month)
        # 遍历工人数据
        for day in range(1, 32):
            working_time = None
            res = df_month[df_month['day'] == day]['工人工日']

            working_time = res.iloc[0] if not res.empty else '/'

            cell_day = dst_sheet.cell(row=row_i, column=day + 1, value=working_time)
            set_style(cell_day)

        row_i += 1

    # 按照每日统计
    dst_sheet.row_dimensions[row_i].height = 20
    table_sumDay = dataframe.groupby(['day'])['工人工日'].sum().reset_index(name='sumday')
    total_time = dataframe['工人工日'].sum()
    total_money = dataframe['劳务费小计'].sum()

    name_t = dst_sheet.cell(row=row_i, column=1, value='合计')
    set_style(name_t)
    cell_total_time = dst_sheet.cell(row=row_i, column=33, value=total_time)
    cell_total_money = dst_sheet.cell(row=row_i, column=35, value=total_money)

    set_style(cell_total_time)
    set_style(cell_total_money)

    for day in range(1, 32):
        sum_day = None
        search = table_sumDay[table_sumDay['day'] == day]['sumday']
        sum_day = search.iloc[0] if not search.empty else 0

        cell_total_day = dst_sheet.cell(row=row_i, column=day + 1, value=sum_day)

        set_style(cell_total_day)

    loc_row = row_i + 1
    row_index = clone_footer(cloneObj, dst_sheet, loc_row)
    return row_index


def create_detail_tables(cloneObj, df, dst_sheet):
    month_group = df.groupby(['month'])

    loc_row = 1
    for month, dataframe in month_group:
        projectName = dataframe.iloc[0, 0]
        year = dataframe.iloc[0, 12]

        month_word = '月份:' + str(year) + '年' + str(month) + '月'

        dataframe = dataframe.reset_index(drop=True)
        offset_index = create_table(projectName, month_word, loc_row, dataframe, cloneObj, dst_sheet)
        loc_row += offset_index + 5


def create_model_sheets(cloneObj, data_obj: DataDeal):
    wb = cloneObj.workbook
    group = data_obj.group_by_team()

    empty_sheet = wb['empty']

    for team, df in group:
        try:
            sheet = wb[team]
        except Exception as err:
            sheet = wb.copy_worksheet(empty_sheet)
            sheet.title = team

        create_detail_tables(cloneObj, df, sheet)


def create_summary_sheet(wb, data_obj: DataDeal):
    """
    汇总
    """
    group = data_obj.sum_money_by_team_month()
    sum_money = data_obj.sum_money_by_team()
    total_money = data_obj.sum_money()

    sheet = wb.create_sheet('总汇')
    # sheet.title = '总汇'
    sheet.column_dimensions['A'].width = 14
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20

    sheet.row_dimensions[1].height = 35
    sheet.row_dimensions[2].height = 30
    # 标题
    cell = sheet.cell(row=1, column=1, value='汇总表')
    cell.font = font16b
    cell.alignment = alignment_center
    sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=4)

    # 列名
    column_name = ['班组', '月份', '金额', '总数']
    for i, name in enumerate(column_name):
        cell = sheet.cell(row=2, column=i + 1, value=name)
        cell.font = font14b
        cell.alignment = alignment_center

    # 主体
    tamp_team_name = None

    # 填充颜色
    color_fun = get_color()
    color = color_fun.__next__()

    for i, row in group.iterrows():

        name = row['班组名称']
        month = row['month']
        money = row['sum']
        all_money = sum_money[sum_money['班组名称'] == name].iloc[0, 1]

        values = [name, month, money, all_money]
        is_merge = False
        # 需要合并单元格
        if tamp_team_name == name:
            is_merge = True
        else:
            tamp_team_name = name
        # 针对不同班组间进行颜色填充
        if not is_merge:
            color = color_fun.__next__()
        for col, val in enumerate(values):
            now_row_index = 3 + int(str(i))
            cell = sheet.cell(row=now_row_index, column=col + 1, value=val)
            cell.fill = color
            cell.font = font12b
            cell.alignment = alignment_center
            cell.border = border_style
            sheet.row_dimensions[now_row_index].height = 30

            if col == 0 or col == 3:
                if is_merge:
                    sheet.merge_cells(start_row=now_row_index - 1, end_row=now_row_index, start_column=col + 1,
                                      end_column=col + 1)

    # 最右侧添加一个总计
    cell = sheet.cell(row=1, column=5, value='总劳务费')
    cell.font = font14b
    cell.fill = fill_棕黄色
    cell.alignment = alignment_center
    cell.border = border_style

    cell = sheet.cell(row=2, column=5, value=total_money)
    cell.font = font14b
    cell.fill = fill_棕黄色
    cell.alignment = alignment_center
    cell.border = border_style


def create_excel(df, save_path, model_path):
    clone_excel = Clone(model_path)
    wb = clone_excel.workbook

    data_obj = DataDeal(df)

    # 使用模板的workbook
    create_summary_sheet(wb, data_obj)

    create_model_sheets(clone_excel, data_obj)
    # 删除模板
    clone_excel.del_model_sheets()
    wb.save(save_path)


if __name__ == '__main__':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    argv = sys.argv
    argv = ['/Users/mac/Downloads/TridentSystem/scripts/python/creatExcel_template_xianchanglaowugz.py',
            'res.xlsx',
            't.json',
            'xianchanglaowugz.xlsx']

    excel_filePath = None
    json_filePath = None
    modelExcel_filePath = None
    if len(argv) > 1:
        try:
            excel_filePath = argv[1]
            json_filePath = argv[2]
            modelExcel_filePath = argv[3]
        except Exception as err:
            print('err:', '模板Excel生成参数不足')

    # 读取 excel json 文件
    if json_filePath:
        with open(json_filePath, encoding='utf8') as f:
            json_args: dict = json.loads(f.read())

            # 读取关键参数
            # 劳务公司代码
            companyCode = json_args.get('劳务公司代码')
            # 项目代码
            projectCode = json_args.get('项目代码')
            # 班组代码
            teamCode = json_args.get('班组代码')
            # 工人代码
            workerCode = json_args.get('工人代码')
            # 筛选时间
            startTime = json_args.get('起始时间')
            endTime = json_args.get('结束时间')

            print(projectCode, teamCode, workerCode, startTime, endTime)

    sql = "select  * from FT256D现场工人出勤统计('%s','%s','%s',%d,'%s','%s')" % (
        teamCode, companyCode, workerCode, projectCode, startTime, endTime)

    # try:
    d = SqlData(sql)
    data: pd.DataFrame = d.data
    if d.length > 0:
        create_excel(data, excel_filePath, modelExcel_filePath)
        print('OK', end='')
    else:
        print('')
    # except Exception as err:
    #     print('',err)
