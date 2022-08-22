from 操作excel.excelApi import sqldata
from 操作excel.excelApi import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils import get_column_letter, column_index_from_string

sql = "select  * from FT256D现场工人出勤统计('*','*','*',10429,'2022-05-01','2022-05-30')"
data: pd.DataFrame = sqldata.Data(sql).data

# 增加一个新列
data['工人工日'] = pd.to_numeric(data['工人工日'])
data['劳务费小计'] = pd.to_numeric(data['劳务费小计'])

data['month'] = data['考勤日期'].astype('str').str.split('-').str.get(1).astype('int')

# 统计 数据准备
group: pd.DataFrame = data.groupby(['工人姓名', 'month'])['劳务费小计'].sum().reset_index(name='sum')

total: pd.DataFrame = group.groupby(['工人姓名'])['sum'].sum().reset_index(name='total')
print(total)
print(total['total'].sum())


def whole_style(sheet, w, h):
    excel_w = max(w)
    excel_h = max(h)

    for w in range(1, excel_w + 1):
        for h in range(1, excel_h + 1):
            if w == 1 and h == 1:
                ...
            else:
                cell = sheet.cell(row=h, column=w)
                cell.border = border_style


def excel():
    ROWS = set()
    COLS = set()

    wb = Workbook()

    sheet = wb.active
    sheet.column_dimensions['A'].width = 14
    sheet.column_dimensions['B'].width = 14
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 25

    sheet.row_dimensions[1].height = 35
    sheet.row_dimensions[2].height = 30

    # 标题
    cell = sheet.cell(row=1, column=1, value='汇总表')
    cell.font = font16p
    cell.alignment = alignment_center
    sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=4)

    # 列名
    column_name = ['姓名', '月份', '金额', '总数']
    for i, name in enumerate(column_name):
        cell = sheet.cell(row=2, column=i + 1, value=name)
        cell.font = font14
        cell.alignment = alignment_center

    # 主体
    for i, row in group.iterrows():

        name = row['工人姓名']
        month = row['month']
        sum_money = row['sum']
        sum_all = total[total['工人姓名'] == name]['total'].iloc[0]

        values = [name, month, sum_money, sum_all]
        for col, val in enumerate(values):
            cell = sheet.cell(row=3 + int(str(i)), column=col + 1, value=val)
            cell.font = font12
            cell.alignment = alignment_center
            sheet.row_dimensions[3 + int(str(i))].height = 30

    wb.save('工资统计.xlsx')


if __name__ == '__main__':
    excel()
