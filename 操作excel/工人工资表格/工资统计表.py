import openpyxl
from 操作excel.excelApi import sqldata
from 操作excel.excelApi import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy
import re

sql = "select  * from FT256D现场工人出勤统计('*','*','*',10429,'2022-05-01','2022-05-30')"
data: pd.DataFrame = sqldata.Data(sql).data

# 增加一个新列
data['工人工日'] = pd.to_numeric(data['工人工日'])
data['劳务费小计'] = pd.to_numeric(data['劳务费小计'])
data['day'] = data['考勤日期'].astype('str').str.split('-').str.get(2).astype('int')

group_class = data.groupby(['班组名称'])


def set_style(cell):
    cell.font = font8
    cell.border = border_style
    cell.alignment = alignment_center


def cloneCell(srcCell, dstCell):
    if srcCell == dstCell:
        return
    dstCell.value = srcCell.value
    if srcCell.has_style:
        dstCell.font = copy(srcCell.font)
        dstCell.border = copy(srcCell.border)
        dstCell.fill = copy(srcCell.fill)
        dstCell.number_format = copy(srcCell.number_format)
        dstCell.protection = copy(srcCell.protection)
        dstCell.alignment = copy(srcCell.alignment)


def clone_footer(scope, src, dst):
    # (row,col)
    # 范围左上角到右下角
    row_s, col_s = scope[0]
    row_e, col_e = scope[1]
    for row in range(1, row_e - row_s + 2):
        for col in range(1, col_e - col_s + 2):
            src_cell = src.cell(row=row, column=col)
            dst_cell = dst.cell(row=row_s + row - 1, column=col_s + col - 1)
            dst.row_dimensions[row_s + row - 1].height = 20
            cloneCell(src_cell, dst_cell)
            # dst.merged_cells


def letter_to_num(p):
    """
    传入 A1 输出 （1，1）
    """
    # pattern = re.compile(r'(\D*)(\d*)')

    letter = re.findall('(\D+)(\d+)', p)[0]
    col = column_index_from_string(letter[0])
    row = int(letter[1])
    return row, col


def deal_mergeInfo(sheet, info, offset):
    """
    模板从（1，1）开始
    """
    print(info)
    row_off = offset[0] - 1
    col_off = offset[1] - 1
    for _ in info:
        m_li = str(_).split(':')
        # 获取点的坐标，对点进行实际的偏移计算
        p1, p2 = letter_to_num(m_li[0]), letter_to_num(m_li[1])

        p1 = p1[0] + row_off, p1[1] + col_off
        p2 = p2[0] + row_off, p2[1] + col_off
        print(p1, p2)
        sheet.merge_cells(start_row=p1[0], start_column=p1[1], end_row=p2[0], end_column=p2[1])


def excel():
    model_wb = load_workbook('工资统计模板.xlsx')
    model_sheet = model_wb['one']
    model_sheet2 = model_wb['two']

    merge_info = model_sheet2.merged_cells

    for sheet_name, df in group_class:
        sheet = model_wb.copy_worksheet(model_sheet)
        sheet.title = sheet_name
        # 主体数据添加
        # 数据处理，提炼出 日期
        df = df.reset_index(drop=True)

        row_i = 5
        # 在对人 进行group
        group_name = df.groupby('工人姓名')
        for name, df_month in group_name:
            cell_c1 = sheet.cell(row=row_i, column=1, value=name)
            set_style(cell_c1)
            sheet.row_dimensions[row_i].height = 20

            time_sum = (df_month['工人工日'] * 100).sum() / 100
            price = df_month.loc[:, '工日单价'].iloc[0]
            total_price = df_month['劳务费小计'].sum()

            cell_AG = sheet.cell(row=row_i, column=33, value=time_sum)
            set_style(cell_AG)

            cell_AH = sheet.cell(row=row_i, column=34, value=price)
            set_style(cell_AH)

            cell_AI = sheet.cell(row=row_i, column=35, value=total_price)
            set_style(cell_AI)

            # print(df_month)
            # 遍历工人数据
            for day in range(1, 32):
                working_time = None
                res = df_month[df_month['day'] == day]['工人工日']

                working_time = res.iloc[0] if not res.empty else '/'

                cell_day = sheet.cell(row=row_i, column=day + 1, value=working_time)
                set_style(cell_day)

            row_i += 1

        # 按照每日统计
        sheet.row_dimensions[row_i].height = 20
        table_sumDay = df.groupby(['day'])['工人工日'].sum().reset_index(name='sumday')
        total_time = df['工人工日'].sum()
        total_money = df['劳务费小计'].sum()

        name_t = sheet.cell(row=row_i, column=1, value='合计')
        set_style(name_t)
        cell_total_time = sheet.cell(row=row_i, column=33, value=total_time)
        cell_total_money = sheet.cell(row=row_i, column=35, value=total_money)

        set_style(cell_total_time)
        set_style(cell_total_money)

        for day in range(1, 32):
            sum_day = None
            search = table_sumDay[table_sumDay['day'] == day]['sumday']
            sum_day = search.iloc[0] if not search.empty else 0

            cell_total_day = sheet.cell(row=row_i, column=day + 1, value=sum_day)

            set_style(cell_total_day)

        clone_footer(scope=[(row_i + 1, 1), (row_i + 4, 35)], src=model_sheet2, dst=sheet)
        deal_mergeInfo(sheet=sheet, info=merge_info, offset=(row_i + 1, 1))

    model_wb.save('res.xlsx')


if __name__ == '__main__':

    excel()
